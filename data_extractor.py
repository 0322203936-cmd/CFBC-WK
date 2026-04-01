"""
data_extractor.py
Centro Floricultor de Baja California
- Hojas WK  → Excel en SharePoint/OneDrive (pandas + requests)
- Hojas PR  → Excel separado en SharePoint  (pandas + requests)
- Hojas MP  → Excel separado en SharePoint  (MANTENIMIENTO)
- Hojas ME  → Excel separado en SharePoint  (MATERIAL DE EMPAQUE)
"""

import re
import requests
import pandas as pd
import openpyxl
from copy import copy
from io import BytesIO

# ─── URLs de SharePoint ───────────────────────────────────────────────────────
# Archivo principal: hojas WK####
SHAREPOINT_URL_WK = (
    "https://pacificafarms-my.sharepoint.com/:x:/g/personal/"
    "anahi_mora_cfbc_co/IQAQCb79SzHtRrTQR71pSNQcASOWqFXyeGGzEhUcT9FRRJ4?e=ClxLCN"
)

# Archivo secundario: hojas PR####, MP####, ME####
SHAREPOINT_URL_PR = (
    "https://pacificafarms-my.sharepoint.com/:x:/g/personal/"
    "jesus_sandoval_cfbc_co/IQCecMwUnigFQa1m-0AYEw-rAenSSKPasiHLi1p2cqtPHkc?e=wpBfv7"
)

# ─── Constantes ───────────────────────────────────────────────────────────────
RANCH_KEYS = ["PROP", "POSCO", "CAMPO", "ISABEL", "HOOPS", "CECILIA", "CHRISTINA", "ALBAHACA"]

CATEGORIAS_ORDEN = [
    "DESINFECCION Y FERTILIZACION",
    "AMPLIACION",
    "CULTIVO TIERRA, CHAROLAS",
    "MATERIAL VEGETAL",
    "PREPARACION DE SUELO",
    "FERTILIZANTES",
    "DESINFECCION / PLAGUICIDAS",
    "MANTENIMIENTO",
    "EXPANSION CECILIA 25",
    "RENOVACION DE SIEMBRA",
    "MATERIAL DE EMPAQUE",
]

SKIP = {"ACUMULADO", "GRAFICOS I-IV", "COMPARATIVO", "DATOS", "HOJA1", "SHEET1"}


# ─── Descarga de Excel desde SharePoint ──────────────────────────────────────
def _descargar_excel(url: str, label: str = "archivo") -> BytesIO | None:
    """
    Descarga un archivo .xlsx desde un link público de SharePoint/OneDrive.
    Agrega el parámetro download=1 necesario para la descarga directa.
    """
    download_url = url.replace("?e=", "?download=1&e=")
    try:
        response = requests.get(download_url, timeout=30)
        response.raise_for_status()
        return BytesIO(response.content)
    except Exception as e:
        print(f"❌ Error descargando {label}: {e}")
        return None


# Alias para compatibilidad con get_sheet_xlsx
def descargar_excel() -> BytesIO | None:
    return _descargar_excel(SHAREPOINT_URL_WK, "Excel WK")


def _leer_hoja(xls: pd.ExcelFile, titulo: str, rango_filas: int = 60,
               rango_cols: int = 35) -> list[list]:
    """
    Lee una hoja del ExcelFile y la retorna como lista de listas.
    Las celdas vacías / NaN se convierten a "".
    """
    try:
        df = pd.read_excel(
            xls,
            sheet_name=titulo,
            header=None,
            nrows=rango_filas,
        ).fillna("")
        if df.shape[1] > rango_cols:
            df = df.iloc[:, :rango_cols]
        return df.values.tolist()
    except Exception as e:
        print(f"   ⚠️  Error leyendo hoja '{titulo}': {e}")
        return []


# ─── Helpers de normalización ─────────────────────────────────────────────────
def norm_ranch(s: str):
    s = str(s).upper().strip()
    if "PROP" in s:                                      return "Prop-RM"
    if "POSCO" in s:                                     return "PosCo-RM"
    if "CAMPO-VI" in s or "CAMPO-IV" in s:               return "Campo-VI"
    if "ALBAHACA" in s:                                  return "Albahaca-RM"
    if "HOOPS" in s:                                     return "HOOPS"
    if "CHRISTINA" in s:                                 return "Christina"
    if "CECILIA 25" in s:                                return "Cecilia 25"
    if "CECILIA" in s:                                   return "Cecilia"
    if "ISABEL" in s:                                    return "Isabela"
    if "CAMPO" in s and "VI" not in s and "IV" not in s: return "Campo-RM"
    return None


def norm_cat(s: str):
    s = str(s).upper().strip()
    if "DESINFECCION" in s and "FERTILIZ" in s:  return "DESINFECCION Y FERTILIZACION"
    if s.startswith("AMPLIACION"):                return "AMPLIACION"
    if "CULTIVO" in s:                            return "CULTIVO TIERRA, CHAROLAS"
    if "MATERIAL VEG" in s:                       return "MATERIAL VEGETAL"
    if "PREPARACION" in s:                        return "PREPARACION DE SUELO"
    if "FERTILIZANTE" in s:                       return "FERTILIZANTES"
    if "SANIDAD" in s or "PLAGUICIDA" in s:       return "DESINFECCION / PLAGUICIDAS"
    if "MANTENIMIENTO" in s:                      return "MANTENIMIENTO"
    if "EXPANSION" in s:                          return "EXPANSION CECILIA 25"
    if "RENOVACION" in s:                         return "RENOVACION DE SIEMBRA"
    if "MATERIAL DE EMP" in s:                    return "MATERIAL DE EMPAQUE"
    if "COSTO DE MAT" in s:                       return "COSTO_STOP"
    if "COSTO DE SERV" in s:                      return "COSTO_STOP"
    if s.startswith("ELECTRICIDAD"):                        return "SV:Electricidad"
    if s.startswith("FLETES Y ACARREOS"):                   return "SV:Fletes y Acarreos"
    if s.startswith("GASTOS DE EXPORTACION"):               return "SV:Gastos de Exportación"
    if s.startswith("CERTIFICADO DE FITOSANITARIO"):        return "SV:Certificado Fitosanitario"
    if s.startswith("TRANSPORTE DE PERSONAL"):              return "SV:Transporte de Personal"
    if s.startswith("COMPRA DE FLOR"):                      return "SV:Compra de Flor a Terceros"
    if s.startswith("COMIDA PARA EL PERSONAL"):             return "SV:Comida para el Personal"
    if s.startswith("RO, TEL") or s.startswith("RO , TEL"): return "SV:RO, TEL, RTA.Alim"
    return None


def sv(v) -> float:
    try:
        if isinstance(v, str):
            v = v.replace("$", "").replace(",", "").strip()
        f = float(v)
        return f if f == f else 0.0
    except (TypeError, ValueError):
        return 0.0


# ─── Parser de hojas PR#### ───────────────────────────────────────────────────
def _parse_pr(rows: list) -> dict:
    RANCH_MAP = {
        'VIV': 'Prop-RM',
        'RAM': 'Campo-RM',
        'ISA': 'Isabela',
        'CHR': 'Christina',
        'CEC': 'Cecilia',
        'C25': 'Cecilia 25',
        'POS': 'PosCo-RM',
        'CAM': 'Campo-RM',
        'ALB': 'Albahaca-RM',
        'HOO': 'HOOPS',
    }
    return _parse_generic(rows, RANCH_MAP)


# ─── Parser de hojas MP#### (MANTENIMIENTO) ───────────────────────────────────
def _parse_mp(rows: list) -> dict:
    RANCH_MAP = {
        'VIV': 'Prop-RM',
        'POS': 'PosCo-RM',
        'RAM': 'Campo-RM',
        'ISA': 'Isabela',
        'CEC': 'Cecilia',
        'C25': 'Cecilia 25',
        'CHR': 'Christina',
    }
    return _parse_generic(rows, RANCH_MAP)


# ─── Parser de hojas ME#### (MATERIAL DE EMPAQUE) ────────────────────────────
def _parse_me(rows: list) -> dict:
    RANCH_MAP = {
        'VIV': 'Prop-RM',
        'POS': 'PosCo-RM',
        'LIM': 'PosCo-RM',
        'RAM': 'Campo-RM',
        'ISA': 'Isabela',
        'CEC': 'Cecilia',
        'C25': 'Cecilia 25',
        'CHR': 'Christina',
        'ALB': 'Albahaca-RM',
        'HOO': 'HOOPS',
    }
    return _parse_generic(rows, RANCH_MAP)


# ─── Parser genérico compartido (PR / MP / ME tienen el mismo formato) ────────
def _parse_generic(rows: list, ranch_map: dict) -> dict:
    """
    Formato común a PR####, MP####, ME####:
      Col 2: UBICACION  (ej: RAMMIPRNN, CECMIPSNF)
      Col 5: PRODUCTO
      Col 7: UNIDADES
      Col 9: GASTO
    Retorna: { rancho: { tipo: [[producto, unidades, gasto, ubicacion], ...] } }
    """
    UBICACION_COL = 2
    PRODUCTO_COL  = 5
    UNIDADES_COL  = 7
    GASTO_COL     = 9

    result = {}
    accum  = {}   # (rancho, tipo, producto, ubicacion) → [u_total, g_total]

    for row in rows:
        if not row or len(row) < 10:
            continue

        ubicacion = str(row[UBICACION_COL]).strip().upper() if len(row) > UBICACION_COL else ''
        ubicacion = re.sub(r'\s+', '', ubicacion)

        if not ubicacion or len(ubicacion) < 6:
            continue
        if not re.match(r'^[A-Z0-9]+$', ubicacion):
            continue

        ranch_code = ubicacion[:3]
        rancho = ranch_map.get(ranch_code)

        if not rancho and ubicacion.startswith('VIV'):
            rancho = 'Prop-RM'

        if not rancho:
            continue

        tipo = 'MIPE' if 'MIP' in ubicacion else 'MIRFE'

        producto = str(row[PRODUCTO_COL]).strip() if len(row) > PRODUCTO_COL else ''
        if not producto or producto.upper() in ('PRODUCTO', 'NOMBRE', ''):
            continue

        unidades = str(row[UNIDADES_COL]).strip() if len(row) > UNIDADES_COL else ''
        try:
            u = float(str(unidades).replace(',', ''))
            unidades = str(int(u)) if u == int(u) else str(round(u, 2))
        except Exception:
            unidades = '0'

        gasto = str(row[GASTO_COL]).strip() if len(row) > GASTO_COL else ''
        try:
            g = float(str(gasto).replace(',', ''))
            gasto = str(round(g, 2))
        except Exception:
            gasto = '0'

        u_f = float(unidades) if unidades else 0.0
        g_f = float(gasto)    if gasto    else 0.0

        key = (rancho, tipo, producto, ubicacion)
        if key in accum:
            accum[key][0] += u_f
            accum[key][1] += g_f
        else:
            accum[key] = [u_f, g_f]

    for (rancho, tipo, producto, ubicacion), (u_tot, g_tot) in accum.items():
        u_str = str(int(u_tot)) if u_tot == int(u_tot) else str(round(u_tot, 2))
        g_str = str(round(g_tot, 2))
        result.setdefault(rancho, {}).setdefault(tipo, []).append([producto, u_str, g_str, ubicacion])

    return result


# ─── Fetch hojas PR / MP / ME desde el segundo Excel de SharePoint ────────────
def _fetch_desde_sharepoint(prefix: str, parser_fn, label: str) -> tuple[dict, dict]:
    """
    Descarga el Excel secundario de SharePoint y extrae todas las hojas
    que coincidan con el patrón  {PREFIX}####  (ej: PR2611, MP2608, ME2610).

    Args:
        prefix:    "PR", "MP" o "ME"
        parser_fn: función que convierte list[list] → dict de ranchos
        label:     nombre legible para logs

    Returns:
        (datos, debug)  con el mismo formato que antes usaban las funciones gspread
    """
    datos = {}
    debug = {f"hojas_{prefix.lower()}_encontradas": []}

    archivo = _descargar_excel(SHAREPOINT_URL_PR, f"Excel {label}")
    if archivo is None:
        print(f"⚠️  No se pudo descargar el archivo para hojas {prefix}")
        return datos, debug

    try:
        xls = pd.ExcelFile(archivo)
    except Exception as e:
        print(f"⚠️  No se pudo abrir el Excel de {label}: {e}")
        return datos, debug

    hojas_encontradas = []
    pat = re.compile(rf'^{prefix}\s*\d{{4}}$', re.IGNORECASE)

    for sname in xls.sheet_names:
        sname = sname.strip()
        if pat.match(sname):
            raw_code = re.sub(rf'{prefix}\s*', '', sname, flags=re.IGNORECASE).strip()
            try:
                code = int(raw_code)
                year = 2000 + (code // 100)
                if 2018 <= year <= 2030:
                    print(f"   ✅ {prefix}{code} encontrada en SharePoint: {sname}")
                    hojas_encontradas.append((sname, code))
                else:
                    print(f"   ❌ {prefix}{code} año {year} fuera de rango")
            except ValueError as e:
                print(f"   ❌ Error código '{raw_code}': {e}")

    debug[f"hojas_{prefix.lower()}_encontradas"] = [t for t, _ in hojas_encontradas]

    if not hojas_encontradas:
        print(f"   ℹ️  No hay hojas {prefix} en el Excel de SharePoint")
        return datos, debug

    for titulo, code in hojas_encontradas:
        vals   = _leer_hoja(xls, titulo, rango_filas=500, rango_cols=11)
        parsed = parser_fn(vals)
        datos[code] = parsed
        debug[f"{prefix}{code}_ranchos"] = list(parsed.keys()) if parsed else []
        print(f"   📦 {prefix}{code} ranchos detectados: {list(parsed.keys())}")

    return datos, debug


# ─── Extractor principal ──────────────────────────────────────────────────────
def extraer_datos(xls: pd.ExcelFile) -> dict:
    all_data       = []
    servicios_data = []

    hojas_validas = []
    pr_hojas      = []

    print("\n" + "=" * 60)
    print("🔍 DETECTANDO HOJAS EN EL EXCEL WK")
    print("=" * 60)

    for sname in xls.sheet_names:
        sname = sname.strip()
        print(f"\n📄 Hoja: '{sname}'")

        if sname.upper() in SKIP:
            print("   ⏭️  SKIP (en lista de exclusión)")
            continue

        pr_match = re.match(r'^PR\s*\d{4}$', sname, re.IGNORECASE)
        if pr_match:
            pr_raw = re.sub(r'PR\s*', '', sname, flags=re.IGNORECASE).strip()
            try:
                pr_code = int(pr_raw)
                pr_year = 2000 + (pr_code // 100)
                if 2018 <= pr_year <= 2030:
                    print("   ✅ PR DETECTADA Y VÁLIDA (en WK Excel)")
                    pr_hojas.append((sname, pr_code))
                    continue
            except ValueError:
                pass

        wk_match = re.match(r'^WK\s*\d{4}$', sname, re.IGNORECASE)
        if wk_match:
            code_raw = re.sub(r"WK\s*", "", sname, flags=re.IGNORECASE).strip()
            try:
                code = int(code_raw)
                year = 2000 + (code // 100)
                if 2018 <= year <= 2030:
                    print("   ✅ WK DETECTADA Y VÁLIDA")
                    hojas_validas.append((sname, code))
                else:
                    print(f"   ❌ Año {year} fuera de rango")
            except ValueError:
                print("   ❌ Error convirtiendo código")
        else:
            if not pr_match:
                print("   ℹ️  No es WK ni PR")

    print("\n" + "=" * 60)
    print("📊 RESUMEN:")
    print(f"   • Hojas WK encontradas: {len(hojas_validas)}")
    print(f"   • Hojas PR en WK Excel: {len(pr_hojas)}")
    print("=" * 60 + "\n")

    if not hojas_validas:
        return {"error": "No se encontraron hojas WK validas."}

    # 2. Leer hojas WK
    batch_data = {}
    for titulo, _ in hojas_validas:
        batch_data[titulo] = _leer_hoja(xls, titulo, rango_filas=120, rango_cols=35)

    # 2b. Leer hojas PR que estén en el Excel WK (fallback)
    productos       = {}
    productos_debug = {"hojas_pr_encontradas": [t for t, _ in pr_hojas]}
    for titulo, pr_code in pr_hojas:
        vals   = _leer_hoja(xls, titulo, rango_filas=500, rango_cols=11)
        parsed = _parse_pr(vals)
        productos[pr_code] = parsed
        productos_debug[f"PR{pr_code}_ranchos"] = list(parsed.keys()) if parsed else []

    # 3. Procesar cada hoja WK
    for titulo, code in hojas_validas:
        raw = batch_data.get(titulo, [])
        if not raw:
            continue

        yy   = code // 100
        ww   = code % 100
        year = 2000 + yy

        max_cols = max((len(r) for r in raw), default=0)
        data     = [r + [""] * (max_cols - len(r)) for r in raw]

        date_range = ""
        for _dr in range(min(8, len(data))):
            for _dc in range(min(5, len(data[_dr]))):
                _v = str(data[_dr][_dc]).strip()
                if _v and " al " in _v.lower() and len(_v) > 8:
                    date_range = _v
                    break
            if date_range:
                break

        exec_idx = -1
        for i, row in enumerate(data):
            if any(isinstance(c, str) and "EJECUCION SEMANAL" in c.upper() for c in row):
                exec_idx = i
                break
        if exec_idx < 0:
            continue

        header_idx = -1
        for i in range(exec_idx - 1, max(0, exec_idx - 6) - 1, -1):
            if any(isinstance(v, str) and any(k in v.upper() for k in RANCH_KEYS) for v in data[i]):
                header_idx = i
                break
        if header_idx < 0:
            continue

        header = data[header_idx]

        total_cols = [j for j, v in enumerate(header)
                      if isinstance(v, str) and v.strip().upper() == "TOTAL"]
        if not total_cols:
            continue
        mxn_total_col = total_cols[0]
        usd_total_col = total_cols[1] if len(total_cols) >= 2 else None

        mxn_ranch_cols, usd_ranch_cols = {}, {}
        for j, v in enumerate(header):
            rn = norm_ranch(str(v)) if v else None
            if not rn:
                continue
            if j < mxn_total_col:
                mxn_ranch_cols[j] = rn
            elif usd_total_col and mxn_total_col < j < usd_total_col:
                mxn_ranch_cols[j] = rn
            elif usd_total_col and j > usd_total_col:
                usd_ranch_cols[j] = rn

        print(f"\n[DEBUG {titulo}]")
        print(f"   exec_idx={exec_idx}, header_idx={header_idx}")
        print(f"   mxn_total_col={mxn_total_col}, usd_total_col={usd_total_col}")
        print(f"   mxn_ranch_cols={mxn_ranch_cols}")
        print(f"   usd_ranch_cols={usd_ranch_cols}")
        hdr_vals = [(j, str(header[j])[:15]) for j in range(len(header)) if str(header[j]).strip()]
        print(f"   header non-empty: {hdr_vals}")

        for i in range(exec_idx + 1, min(exec_idx + 120, len(data))):
            row   = data[i]
            label = next((str(row[c]).strip() for c in range(5)
                          if c < len(row) and row[c] and len(str(row[c]).strip()) > 3), None)
            if not label:
                continue

            cat = norm_cat(label)
            if not cat:
                continue
            if cat == "COSTO_STOP":
                continue

            mxn_ranches = {rn: sv(row[j]) for j, rn in mxn_ranch_cols.items() if j < len(row)}
            usd_ranches = {rn: sv(row[j]) for j, rn in usd_ranch_cols.items() if j < len(row)}

            if cat.startswith("SV:"):
                print(f"   [SV] fila={i} label='{label[:30]}' cat='{cat}' mxn_ranches={mxn_ranches}")
                servicios_data.append({
                    "semana":      code,
                    "year":        year,
                    "week":        ww,
                    "date_range":  date_range,
                    "subcat":      cat[3:],
                    "mxn_total":   round(sv(row[mxn_total_col]) if mxn_total_col < len(row) else 0, 2),
                    "usd_total":   round(sv(row[usd_total_col]) if usd_total_col and usd_total_col < len(row) else 0, 2),
                    "mxn_ranches": mxn_ranches,
                    "usd_ranches": usd_ranches,
                })
            else:
                all_data.append({
                    "semana":      code,
                    "year":        year,
                    "week":        ww,
                    "date_range":  date_range,
                    "categoria":   cat,
                    "mxn_total":   round(sv(row[mxn_total_col]) if mxn_total_col < len(row) else 0, 2),
                    "usd_total":   round(sv(row[usd_total_col]) if usd_total_col and usd_total_col < len(row) else 0, 2),
                    "mxn_ranches": mxn_ranches,
                    "usd_ranches": usd_ranches,
                })

    print(f"\n✅ servicios_data: {len(servicios_data)} registros encontrados")
    if servicios_data:
        print(f"   subcats: {list({r['subcat'] for r in servicios_data})}")

    cats_found = {r["categoria"] for r in all_data}
    cats  = [c for c in CATEGORIAS_ORDEN if c in cats_found]
    years = sorted({r["year"] for r in all_data})

    ranches_seen: set = set()
    for r in all_data:
        ranches_seen.update(r["mxn_ranches"])
        ranches_seen.update(r["usd_ranches"])
    ranches = sorted(ranches_seen)

    summary: dict = {cat: {yr: {"usd": 0.0, "mxn": 0.0, "ranches": {}, "ranches_mxn": {}}
                            for yr in years} for cat in cats}
    for r in all_data:
        s = summary.get(r["categoria"], {}).get(r["year"])
        if not s:
            continue
        s["usd"] += r["usd_total"]
        s["mxn"] += r["mxn_total"]
        for rn, v in r["usd_ranches"].items():
            s["ranches"][rn] = round(s["ranches"].get(rn, 0) + v, 2)
        for rn, v in r["mxn_ranches"].items():
            s["ranches_mxn"][rn] = round(s["ranches_mxn"].get(rn, 0) + v, 2)
    for cat in cats:
        for yr in years:
            d = summary[cat][yr]
            d["usd"] = round(d["usd"], 2)
            d["mxn"] = round(d["mxn"], 2)

    weeks_per_year: dict = {}
    week_date_ranges: dict = {}
    for r in all_data:
        weeks_per_year.setdefault(r["year"], set()).add(r["week"])
        key = f"{r['year']}-{r['week']}"
        if key not in week_date_ranges and r.get("date_range"):
            week_date_ranges[key] = r["date_range"]
    weeks_per_year = {yr: sorted(wks) for yr, wks in weeks_per_year.items()}

    return {
        "years":            years,
        "categories":       cats,
        "ranches":          ranches,
        "summary":          summary,
        "weeks_per_year":   weeks_per_year,
        "week_date_ranges": week_date_ranges,
        "weekly_detail":    all_data,
        "productos":        productos,
        "productos_debug":  productos_debug,
        "servicios_data":   servicios_data,
    }


# ─── Punto de entrada público ─────────────────────────────────────────────────
def get_datos() -> dict:
    """
    - Hojas WK  → Excel principal en SharePoint
    - Hojas PR  → Excel secundario en SharePoint
    - Hojas MP  → Excel secundario en SharePoint (MANTENIMIENTO)
    - Hojas ME  → Excel secundario en SharePoint (MATERIAL DE EMPAQUE)
    """
    # 1. Descargar y leer Excel WK
    archivo = _descargar_excel(SHAREPOINT_URL_WK, "Excel WK")
    if archivo is None:
        return {"error": "No se pudo descargar el archivo WK de SharePoint."}

    try:
        xls = pd.ExcelFile(archivo)
    except Exception as e:
        return {"error": f"No se pudo abrir el Excel WK: {e}"}

    resultado = extraer_datos(xls)

    if "error" not in resultado:
        # 2. Leer PR desde Excel secundario de SharePoint
        print("\n" + "=" * 60)
        print("🔍 LEYENDO HOJAS PR DESDE SHAREPOINT")
        print("=" * 60)
        productos, productos_debug = _fetch_desde_sharepoint("PR", _parse_pr, "PR")
        # Merge con cualquier PR que ya estuviera en el Excel WK
        resultado["productos"].update(productos)
        resultado["productos_debug"].update(productos_debug)

        # 3. Leer MP desde Excel secundario de SharePoint (MANTENIMIENTO)
        print("\n" + "=" * 60)
        print("🔍 LEYENDO HOJAS MP DESDE SHAREPOINT (MANTENIMIENTO)")
        print("=" * 60)
        productos_mp, productos_mp_debug = _fetch_desde_sharepoint("MP", _parse_mp, "MP")
        resultado["productos_mp"]       = productos_mp
        resultado["productos_mp_debug"] = productos_mp_debug

        # 4. Leer ME desde Excel secundario de SharePoint (MATERIAL DE EMPAQUE)
        print("\n" + "=" * 60)
        print("🔍 LEYENDO HOJAS ME DESDE SHAREPOINT (MATERIAL DE EMPAQUE)")
        print("=" * 60)
        productos_me, productos_me_debug = _fetch_desde_sharepoint("ME", _parse_me, "ME")
        resultado["productos_me"]       = productos_me
        resultado["productos_me_debug"] = productos_me_debug

    return resultado


# --- Construir hoja WK en blanco con estructura fija ---
def _construir_hoja_wk(ws, nombre_hoja: str):
    """
    Escribe la estructura completa de una hoja WK#### desde cero.
    Todos los valores de datos quedan en 0 / vacíos para ser llenados manualmente.
    """
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers)
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

    # ── Helpers de estilo ────────────────────────────────────────────────
    def _f(bold=False, size=10, color="000000", name="Calibri"):
        return Font(bold=bold, size=size, color=color, name=name)

    def _fill(hex_color):
        if not hex_color or hex_color in ("00000000", "FFFFFFFF", ""):
            return PatternFill(fill_type=None)
        c = hex_color.lstrip("FF") if len(hex_color) == 8 else hex_color
        return PatternFill("solid", fgColor=c)

    def _al(h="general", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    thin = Side(style="thin")
    medium = Side(style="medium")

    def _border(left=None, right=None, top=None, bottom=None):
        return Border(left=left, right=right, top=top, bottom=bottom)

    # Fills frecuentes
    fill_green   = _fill("CCE5CC")   # verde claro (USD)
    fill_blue    = _fill("DAE3F3")   # azul claro  (encabezado semana)
    fill_lime    = _fill("C5E0B4")   # verde lima  (código semana)
    fill_orange  = _fill("FFCC99")   # naranja     (totales subtotales)
    fill_yellow  = _fill("FFFFCC")   # amarillo    (producción)
    fill_white   = PatternFill(fill_type=None)

    def _set(cell, value=None, bold=False, fill=None, align_h="general",
             num_fmt=None, font_color="000000"):
        cell.value = value
        cell.font  = _f(bold=bold, color=font_color)
        if fill:
            cell.fill = fill
        cell.alignment = _al(h=align_h)
        if num_fmt:
            cell.number_format = num_fmt

    # ── Ancho de columnas ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 69.4
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 11
    ws.column_dimensions["F"].width = 11
    ws.column_dimensions["G"].width = 11
    ws.column_dimensions["H"].width = 11
    ws.column_dimensions["I"].width = 11
    ws.column_dimensions["J"].width = 11
    ws.column_dimensions["K"].width = 3
    for col in ("L", "M", "N", "O", "P", "Q", "R", "S"):
        ws.column_dimensions[col].width = 11

    # ── Fila 1: Empresa ──────────────────────────────────────────────────
    ws["B1"] = "CENTRO FLORICULTOR DE BAJA CALIFORNIA, S.A. DE C.V. "
    ws["B1"].font = _f(bold=True)

    # ── Fila 2: Encabezado semana ────────────────────────────────────────
    ws["B2"] = "SEMANA DE CALCULO - Mexico"
    ws["B2"].font = _f(bold=True)
    ws["B2"].fill = _fill("DAE3F3")
    ws["B2"].alignment = _al("center")

    # ── Fila 3: Código semana + tipo de cambio ───────────────────────────
    ws["B3"] = nombre_hoja                # ej: WK2613
    ws["B3"].font = _f(bold=True)
    ws["B3"].fill = _fill("C5E0B4")
    ws["B3"].alignment = _al("center")
    ws["C3"] = 19
    ws["C3"].font = _f(bold=True)
    ws["D3"] = " tipo de cambio"
    ws["D3"].font = _f(bold=True)
    ws["L3"] = 19
    ws["L3"].font = _f(bold=True)
    ws["M3"] = "  tipo de cambio "
    ws["M3"].font = _f(bold=True)

    # ── Fila 4: Rango de fechas ──────────────────────────────────────────
    ws["B4"] = "Del ___ al ___ de ________ 20__"
    ws["B4"].alignment = _al("center")
    ws.row_dimensions[4].height = 15

    # ── Fila 5: Etiquetas MXN / USD ─────────────────────────────────────
    ws.merge_cells("C5:J5")
    ws["C5"] = "(MXN) Pesos Mexicanos"
    ws["C5"].alignment = _al("center")
    ws.merge_cells("L5:R5")
    ws["L5"] = "US Dollars"
    ws["L5"].fill = fill_green
    ws["L5"].alignment = _al("center")

    # ── Fila 6: TOTAL FINCA ──────────────────────────────────────────────
    ws["B6"] = "TOTAL FINCA"
    ws["B6"].alignment = _al("center")
    ws.row_dimensions[6].height = 26.4

    # ── Fila 7: Encabezados de ranchos ───────────────────────────────────
    headers_mxn = ["TOTAL", "Prop-RM", "PosCo-RM", "Campo -RM",
                   "Isabela", "Christina", "Cecilia", "Cecilia 25"]
    headers_usd = ["TOTAL", "Prop-RM", "PosCo-RM", "Campo -RM",
                   "ISABELA", "Christina", "CECILIA", "CECILIA 25"]
    for i, h in enumerate(headers_mxn):
        col = chr(ord("C") + i)
        ws[f"{col}7"] = h
        ws[f"{col}7"].alignment = _al("center")
    usd_cols = ["L", "M", "N", "O", "P", "Q", "R", "S"]
    for i, h in enumerate(headers_usd):
        ws[f"{usd_cols[i]}7"] = h
        ws[f"{usd_cols[i]}7"].fill = fill_green
        ws[f"{usd_cols[i]}7"].alignment = _al("center")

    # ── Fila 8: SEMANAL ──────────────────────────────────────────────────
    ws["B7"] = "Produccion"
    ws["C8"] = "SEMANAL "
    ws["C8"].alignment = _al("center")
    ws["L8"] = '"WEEKLY"'
    ws["L8"].fill = fill_green
    ws["L8"].alignment = _al("center")

    # ── Fila 9: EJECUCION SEMANAL + factores ─────────────────────────────
    ws["B9"] = "EJECUCION SEMANAL"
    ws["B9"].font = _f(bold=True)
    ws["B9"].alignment = _al("center")
    for col in ["D", "E", "F", "G", "H", "I", "J"]:
        ws[f"{col}9"] = 1
        ws[f"{col}9"].alignment = _al("center")

    # ── Filas 10-20: Categorías de materiales ────────────────────────────
    categorias = [
        (10, "DESINFECCION Y FERTILIZACION"),
        (11, "AMPLIACION "),
        (12, "CULTIVO TIERRA, CHAROLAS"),
        (13, "MATERIAL VEGETAL"),
        (14, "PREPARACION DE SUELO"),
        (15, "FERTILIZANTES (Manejo Integrado de Riego y Fertilización) "),
        (16, "DESINFECCION / PLAGUICIDAS (Manejo Integrado de Plagas y Enfermedades)"),
        (17, "MANTENIMIENTO"),
        (18, "EXPANSION CECILIA 25"),
        (19, "RENOVACION DE SIEMBRA"),
        (20, "MATERIAL DE EMPAQUE"),
    ]
    for row, label in categorias:
        ws[f"B{row}"] = label
        ws[f"B{row}"].alignment = _al("left")
        ws[f"C{row}"] = 0
        ws[f"C{row}"].font = _f(bold=True)
        ws[f"C{row}"].number_format = '#,##0;-#,##0;"-   "'
        ws[f"L{row}"] = 0
        ws[f"L{row}"].font = _f(bold=True)
        ws[f"L{row}"].fill = fill_green
        ws[f"L{row}"].number_format = '#,##0;-#,##0;"-   "'
        for uc in ["M", "N", "O", "P", "Q", "R", "S"]:
            ws[f"{uc}{row}"] = 0
            ws[f"{uc}{row}"].fill = fill_green
            ws[f"{uc}{row}"].alignment = _al("center")
            ws[f"{uc}{row}"].number_format = '#,##0;-#,##0;" -   "'
        for dc in ["D", "E", "F", "G", "H", "I", "J"]:
            ws[f"{dc}{row}"] = 0
            ws[f"{dc}{row}"].number_format = '#,##0;-#,##0;"-   "'

    # ── Fila 22: COSTO DE MATERIALES (subtotal) ──────────────────────────
    ws["B22"] = "COSTO DE MATERIALES"
    ws["B22"].font = _f(bold=True)
    data_cols_mxn = ["C", "D", "E", "F", "G", "H", "I", "J"]
    data_cols_usd = ["L", "M", "N", "O", "P", "Q", "R", "S"]
    for col in data_cols_mxn:
        ws[f"{col}22"] = 0
        ws[f"{col}22"].font = _f(bold=True)
        ws[f"{col}22"].fill = fill_orange
        ws[f"{col}22"].alignment = _al("center")
        ws[f"{col}22"].number_format = '#,##0;-#,##0;"-   "'
    for col in data_cols_usd:
        ws[f"{col}22"] = 0
        ws[f"{col}22"].font = _f(bold=True)
        ws[f"{col}22"].fill = fill_orange
        ws[f"{col}22"].alignment = _al("center")
        ws[f"{col}22"].number_format = '#,##0;-#,##0;"-   "'

    # ── Filas 24-59: Nóminas y cargas sociales ───────────────────────────
    nominas = [
        (24, "NOMINA ADMON Oficina, Jefes de Finca, Ingenieros"),
        (25, "HORAS EXTR. DOM. Y FESTIVOS"),
        (26, "BONOS ASISIT, PUNTAULIDAD Y DESPENSA"),
        (27, "NOMINA PRODUCCION "),
        (28, "HORAS EXTR. DOM. Y FEST."),
        (29, "BONOS ASISIT, PUNT. Y DESP."),
        (30, "NOMINA PRODUCCION CORTE"),
        (31, "HORAS EXTR. DOM. Y FESTIVOS CORTE"),
        (32, "BONOS ASISIT, PUNTAULIDAD Y DESP. CORTE"),
        (33, "NOMINA PRODUCCION TRANSPLANTE"),
        (34, "HORAS EXTR. DOM. Y FEST. TRANSPLANTE"),
        (35, "BONOS ASISIT, PUNT. Y DESP. TRANSPLANTE"),
        (36, "NOMINA PRODUCCION MANEJO PLANTA"),
        (37, "HORAS EXTR. DOM. Y FEST. MANEJO PLANTA"),
        (38, "BONOS ASISIT, PUNT. Y DESP. MANEJO PLANTA"),
        (39, "NOMINA  HOOPS"),
        (40, "HORAS EXTR. DOM. Y FEST. HOOPS"),
        (41, "BONOS ASISIT, PUNT. Y DESP.HOOPS"),
        (42, "NOMINA  (MIPE,MIRFE,)"),
        (43, "HORAS EXTR. DOM. Y FEST. (MIPE,MIRFE)"),
        (44, "BONOS ASISIT, PUNT. Y DESP.(MIPE,MIRFE)"),
        (45, "NOMINA OPERATIVOS (TRACTORES, CAMEROS)"),
        (46, "HORAS EXTR. DOM. Y FEST. (TRACTORES, CAMEROS)"),
        (47, "BONOS ASISIT, PUNT. Y DESP. (TRACTORES, CAMEROS)"),
        (48, "NOMINA OPERATIVOS (CHOFER)"),
        (49, "HORAS EXTR. DOM. Y FEST. (CHOFER)"),
        (50, "BONOS ASISIT, PUNT. Y DESP. (CHOFER)"),
        (51, "NOMINA OPERATIVOS (VELADORES)"),
        (52, "HORAS EXTR. DOM. Y FEST. (VELADORES)"),
        (53, "BONOS ASISIT, PUNT. Y DESP. (VELADORES)"),
        (54, "NOMINA OPERATIVOS (SOLDADOR)"),
        (55, "HORAS EXTR. DOM. Y FEST. (SOLDADOR)"),
        (56, "BONOS ASISIT, PUNT. Y DESP. (SOLDADOR)"),
        (57, "NOMINA PRODUCCION Contratista y comisiones"),
        (58, "IMSS , INFONAVIT RCV"),
        (59, "1.8% al estado (1.2% tasa efectiva)"),
    ]
    for row, label in nominas:
        ws[f"B{row}"] = label
        ws[f"B{row}"].alignment = _al("left")
        ws[f"C{row}"] = 0
        ws[f"C{row}"].font = _f(bold=True)
        ws[f"C{row}"].number_format = '#,##0;-#,##0;"-   "'
        ws[f"L{row}"] = 0
        ws[f"L{row}"].font = _f(bold=True)
        ws[f"L{row}"].fill = fill_green
        ws[f"L{row}"].number_format = '#,##0;-#,##0;"-   "'
        for uc in ["M", "N", "O", "P", "Q", "R", "S"]:
            ws[f"{uc}{row}"] = 0
            ws[f"{uc}{row}"].fill = fill_green
            ws[f"{uc}{row}"].alignment = _al("center")
            ws[f"{uc}{row}"].number_format = '#,##0;-#,##0;" -   "'
        for dc in ["D", "E", "F", "G", "H", "I", "J"]:
            ws[f"{dc}{row}"] = 0
            ws[f"{dc}{row}"].number_format = '#,##0;-#,##0;"-   "'

    # ── Fila 61: COSTO DE MANO DE OBRA ───────────────────────────────────
    ws["B61"] = "COSTO DE MANO DE OBRA"
    ws["B61"].font = _f(bold=True)
    for col in data_cols_mxn:
        ws[f"{col}61"] = 0
        ws[f"{col}61"].font = _f(bold=True)
        ws[f"{col}61"].fill = fill_orange
        ws[f"{col}61"].number_format = '#,##0;-#,##0;"-   "'
    for col in data_cols_usd:
        ws[f"{col}61"] = 0
        ws[f"{col}61"].font = _f(bold=True)
        ws[f"{col}61"].fill = fill_orange
        ws[f"{col}61"].number_format = '#,##0;-#,##0;"-   "'

    # ── Filas 63-70: Servicios ────────────────────────────────────────────
    servicios = [
        (63, "ELECTRICIDAD"),
        (64, "FLETES Y ACARREOS (Flete aduana)"),
        (65, "GASTOS DE EXPORTACION "),
        (66, "CERTIFICADO DE FITOSANITARIOS"),
        (67, "Transporte de personal"),
        (68, "COMPRA DE FLOR A TERCEROS"),
        (69, "COMIDA PARA EL PERSONAL"),
        (70, "RO, TEL, RTA.ALIM."),
    ]
    for row, label in servicios:
        ws[f"B{row}"] = label
        ws[f"B{row}"].alignment = _al("left")
        ws[f"C{row}"] = 0
        ws[f"C{row}"].font = _f(bold=True)
        ws[f"C{row}"].number_format = '#,##0;-#,##0;"-   "'
        ws[f"L{row}"] = 0
        ws[f"L{row}"].font = _f(bold=True)
        ws[f"L{row}"].fill = fill_green
        ws[f"L{row}"].number_format = '#,##0;-#,##0;"-   "'
        for uc in ["M", "N", "O", "P", "Q", "R", "S"]:
            ws[f"{uc}{row}"] = 0
            ws[f"{uc}{row}"].fill = fill_green
            ws[f"{uc}{row}"].alignment = _al("center")
            ws[f"{uc}{row}"].number_format = '#,##0;-#,##0;" -   "'
        for dc in ["D", "E", "F", "G", "H", "I", "J"]:
            ws[f"{dc}{row}"] = 0
            ws[f"{dc}{row}"].number_format = '#,##0;-#,##0;"-   "'

    # ── Fila 72: COSTO DE SERVICIOS ───────────────────────────────────────
    ws["B72"] = "COSTO DE SERVICIOS"
    ws["B72"].font = _f(bold=True)
    for col in data_cols_mxn:
        ws[f"{col}72"] = 0
        ws[f"{col}72"].font = _f(bold=True)
        ws[f"{col}72"].fill = fill_orange
        ws[f"{col}72"].number_format = '#,##0;-#,##0;"-   "'
    for col in data_cols_usd:
        ws[f"{col}72"] = 0
        ws[f"{col}72"].font = _f(bold=True)
        ws[f"{col}72"].fill = fill_orange
        ws[f"{col}72"].number_format = '#,##0;-#,##0;"-   "'

    # ── Fila 74: COSTO DE PRODUCCION Y VENTAS ────────────────────────────
    ws["B74"] = "COSTO DE PRODUCCION Y VENTAS"
    ws["B74"].font = _f(bold=True)
    ws.row_dimensions[74].height = 15
    for col in data_cols_mxn:
        ws[f"{col}74"] = 0
        ws[f"{col}74"].font = _f(bold=True)
        ws[f"{col}74"].number_format = '#,##0;-#,##0;"-   "'
    for col in data_cols_usd:
        ws[f"{col}74"] = 0
        ws[f"{col}74"].font = _f(bold=True)
        ws[f"{col}74"].fill = fill_green
        ws[f"{col}74"].number_format = '#,##0;-#,##0;"-   "'

    # ── Filas 76-92: Producción (tallos, charolas, hectáreas) ─────────────
    produccion = [
        (76, "CAJAS PROCESADAS TOTALES"),
        (77, "INVENTARIO INICIAL"),
        (78, "TALLOS COSECHADOS"),
        (79, "TALLOS DESECHADOS"),
        (80, "TALLOS DESECHADOS sf"),
        (81, "TALLOS COMPRADOS"),
        (82, "TALLOS EN BOUQUETS O PROCESADOS"),
        (83, "TALLOS DESPACHADOS"),
        (84, "LIBRAS DESPACHADAS ALBAHACA"),
        (85, "TALLOS muestra"),
        (86, "INVENTARIO FINAL"),
        (87, "TALLOS PROCESADOS TOTALES"),
        (88, " CHAROLAS SEMBRADAS *288 PLUGS ="),
        (89, " NUMERO DE CHAROLAS SEMBRADAS "),
        (90, " NUMERO DE ESQUEJES SEMBRADOS"),
        (91, " METROS DE SIEMBRA"),
        (92, " HECTAREAS EN SIEMBRA"),
    ]
    for row, label in produccion:
        ws[f"B{row}"] = label
        ws[f"B{row}"].alignment = _al("left")
        ws[f"C{row}"] = 0
        ws[f"C{row}"].font = _f(bold=True)
        ws[f"L{row}"] = 0
        ws[f"L{row}"].font = _f(bold=True)
        ws[f"L{row}"].fill = fill_yellow
        for uc in ["M", "N", "O", "P", "Q", "R", "S"]:
            ws[f"{uc}{row}"] = 0
            ws[f"{uc}{row}"].fill = fill_yellow
            ws[f"{uc}{row}"].alignment = _al("center")
        for dc in ["D", "E", "F", "G", "H", "I", "J"]:
            ws[f"{dc}{row}"] = 0

    ws.row_dimensions[92].height = 15

    # ── Fila 93: Indicadores ──────────────────────────────────────────────
    ws["B93"] = "<<< INDICADORES"
    ws["B93"].font = _f(bold=True)
    ws.row_dimensions[93].height = 15

    # ── Filas 94-121: Costos unitarios e indicadores ──────────────────────
    ws["B94"] = "COSTOS UNITARIOS"
    ws["B94"].font = _f(bold=True)
    ws["B95"] = "$ / Tallo Procesado"
    ws["B95"].font = _f(bold=True)
    ws["B96"] = "COSTOS UNITARIOS"
    ws["B96"].font = _f(bold=True)
    ws["B97"] = "$ / Libras Procesadas"
    ws["B97"].font = _f(bold=True)

    cu_rows = [
        (98,  "Materiales"),
        (99,  "Mano de Obra"),
        (100, "Servicios (Fletes)"),
        (101, "Costo de Produccion y Ventas"),
        (103, "Material de Empaque / Tallo"),
        (105, "Sanidad Vegetal / Tallo"),
        (106, "Fertlizacion / Tallo"),
        (108, "Mano de Obra Prod / Tallo"),
    ]
    for row, label in cu_rows:
        ws[f"B{row}"] = label
        bold = row == 101
        ws[f"B{row}"].font = _f(bold=bold)
        if bold:
            ws[f"B{row}"].fill = fill_orange
        ws[f"C{row}"] = 0
        ws[f"C{row}"].font = _f(bold=bold)
        ws[f"L{row}"] = 0
        ws[f"L{row}"].font = _f(bold=bold)
        ws[f"L{row}"].fill = fill_green

    ws.row_dimensions[108].height = 15
    ws.row_dimensions[109].height = 15

    ws["B110"] = "$ / Hectarea"
    ws["B110"].font = _f(bold=True)

    ha_rows = [
        (111, "Materiales"),
        (112, "Mano de Obra"),
        (113, "Servicios (Fletes)"),
        (114, "Costo de Produccion y Ventas"),
    ]
    for row, label in ha_rows:
        ws[f"B{row}"] = label
        ws[f"C{row}"] = 0
        ws[f"C{row}"].font = _f(bold=True)
        ws[f"L{row}"] = 0
        ws[f"L{row}"].font = _f(bold=True)
        ws[f"L{row}"].fill = fill_yellow

    ws.row_dimensions[121].height = 15

    ws["B121"] = "Mano de Obra Prod / Ha"
    ws["B121"].font = _f(bold=True)
    ws["C121"] = 0
    ws["C121"].font = _f(bold=True)
    ws["L121"] = 0
    ws["L121"].font = _f(bold=True)
    ws["L121"].fill = fill_yellow

    # ── Fila 124: KPIs ────────────────────────────────────────────────────
    ws["B124"] = "KPI's "
    ws["B124"].font = _f(bold=True)

    # ── Fila 125: Proyectos de inversión ──────────────────────────────────
    ws["B125"] = "Proyectos de inversión"
    ws["B125"].font = _f(bold=True, color="008000")
    ws["L125"] = "Total Weekly"
    ws["L125"].font = _f(bold=True, color="008000")
    ws["L125"].alignment = _al("center")

    proyectos = [
        (126, "Sistema de riego (Ramona)"),
        (127, "Sistema de riego (Isabella)"),
        (128, "Caseta (Isabella)"),
        (129, "Sistema de ventilacion"),
        (130, "Sistema de tratamiento de aguas residuales (Isabella)"),
        (131, "Arcos para invernaderos "),
        (132, "proyecto luz"),
        (133, "Construcción de Almacén (Ramona) "),
        (134, "Construcción de Almacén (Isabela) "),
        (135, "Carritos"),
        (136, "Maquinaria "),
        (137, "Chiller"),
        (138, "Cuarto frio"),
        (139, "veronicas"),
    ]
    for row, label in proyectos:
        ws[f"B{row}"] = label
        ws[f"C{row}"] = 0
        ws[f"C{row}"].number_format = '"$"#,##0;-"$"#,##0;"$-   "'
        ws[f"L{row}"] = 0
        ws[f"L{row}"].number_format = '"$"#,##0;-"$"#,##0;" $-   "'
        for uc in ["M", "N", "O", "P", "Q", "R", "S"]:
            ws[f"{uc}{row}"] = 0

    ws["B140"] = "Total "
    ws["B140"].font = _f(bold=True)
    ws["C140"] = 0
    ws["C140"].number_format = '" $"#,##0;-" $"#,##0;" $-   "'
    ws["L140"] = 0
    ws["L140"].number_format = '" $"#,##0;-" $"#,##0;" $-   "'

    # ── Fila 143: Logística ───────────────────────────────────────────────
    ws["B143"] = "Logística "
    ws["B143"].font = _f(bold=True, color="008000")
    ws["L143"] = "Total Weekly"
    ws["L143"].font = _f(bold=True, color="008000")
    ws["L143"].alignment = _al("center")
    ws["N143"] = "PosCo-RM"
    ws["N143"].font = _f(bold=True, color="008000")
    ws["N143"].alignment = _al("center")

    logistica = [
        (144, "Número de camiones despachados "),
        (145, "Número de tarimas despachadas (montadas al camión)"),
        (146, "Número de cajas despachadas"),
        (147, "Número de Pies cúbicos de cajas despachadas "),
        (148, "Número de Pies cubicos promedio / camión despachado "),
        (149, "Capacidad en pies cúbicos por camión "),
        (150, "Rendimiento promedio por camión "),
    ]
    for row, label in logistica:
        ws[f"B{row}"] = label
        ws[f"C{row}"] = 0
        ws[f"L{row}"] = 0
        ws[f"N{row}"] = 0
        ws[f"N{row}"].font = _f(bold=True)

    # ── Filas 152-172: KPIs de flete y material de empaque ────────────────
    kpi_groups = [
        (152, "Costo incurrido por flete, gtos expo, fitosanitarios"),
        (153, "Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
        (154, "Número de Camiones despachados "),
        (156, "Costo incurrido promedio flete, gtos expo, fitosanitarios / pie cúbico"),
        (157, "Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
        (158, "Número de Pies cúbicos de cajas despachadas"),
        (160, "Costo incurrido flete, gtos expo, fitosanitarios / cajas despachadas"),
        (161, "Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
        (162, "Número de cajas despachadas"),
    ]
    for row, label in kpi_groups:
        ws[f"B{row}"] = label
        ws[f"C{row}"] = 0
        ws[f"L{row}"] = 0
        ws[f"N{row}"] = 0
        ws[f"N{row}"].font = _f(bold=True)

    ws["B165"] = "Material de empaque / Caja"
    ws["B165"].font = _f(bold=True, color="008000")

    me_rows = [
        (166, "Costo incurrido en Material de empaque / pie cúbico"),
        (167, "Costo incurrido en Material de empaque (USD)"),
        (168, "Número de Pies cúbicos de cajas despachadas"),
        (170, "Costo incurrido en Material de empaque / cajas despachadas"),
        (171, "Costo incurrido en Material de empaque (USD)"),
        (172, "Número de cajas despachadas"),
    ]
    for row, label in me_rows:
        ws[f"B{row}"] = label
        ws[f"C{row}"] = 0
        ws[f"L{row}"] = 0
        ws[f"N{row}"] = 0
        ws[f"N{row}"].font = _f(bold=True)

    # ── Merged cells ─────────────────────────────────────────────────────
    merges = [
        "C5:J5", "L5:R5",
        "C153:C154", "L153:L154",
        "C157:C158", "L157:L158",
        "C161:C162", "L161:L162",
        "C167:C168", "L167:L168",
        "C171:C172", "L171:L172",
    ]
    for m in merges:
        try:
            ws.merge_cells(m)
        except Exception:
            pass


# --- Crear nueva hoja WK en SharePoint via Microsoft Graph API (con sesión) ---
def crear_hoja_wk(nombre_hoja: str, tenant_id: str, client_id: str, client_secret: str) -> dict:
    """
    Crea una nueva hoja WK#### desde cero usando una sesión de workbook de Graph API.
    Funciona aunque el archivo esté abierto por otros usuarios (no requiere lock).
    Escribe todas las celdas directamente via API, sin subir el archivo completo.
    Requiere Files.ReadWrite en la App Registration de Azure AD.
    """
    import base64 as _b64
    import time

    # ── Helper: construir lista plana de celdas { address, value } ────────
    def _celdas_de_la_hoja(nombre):
        """Devuelve lista de dicts con address y valor para escribir via Graph API."""
        celdas = []

        def c(addr, val):
            celdas.append({"address": addr, "value": val})

        # Encabezados
        c("B1", "CENTRO FLORICULTOR DE BAJA CALIFORNIA, S.A. DE C.V. ")
        c("B2", "SEMANA DE CALCULO - Mexico")
        c("B3", nombre)
        c("C3", 19); c("D3", " tipo de cambio")
        c("L3", 19); c("M3", "  tipo de cambio ")
        c("B4", "Del ___ al ___ de ________ 20__")
        c("C5", "(MXN) Pesos Mexicanos")
        c("L5", "US Dollars")
        c("B6", "TOTAL FINCA")
        # Fila 7: ranchos MXN
        for col, h in zip(["C","D","E","F","G","H","I","J"],
                          ["TOTAL","Prop-RM","PosCo-RM","Campo -RM","Isabela","Christina","Cecilia","Cecilia 25"]):
            c(f"{col}7", h)
        # Fila 7: ranchos USD
        for col, h in zip(["L","M","N","O","P","Q","R","S"],
                          ["TOTAL","Prop-RM","PosCo-RM","Campo -RM","ISABELA","Christina","CECILIA","CECILIA 25"]):
            c(f"{col}7", h)
        c("B7", "Produccion"); c("C8", "SEMANAL "); c("L8", '"WEEKLY"')
        c("B9", "EJECUCION SEMANAL")
        for col in ["D","E","F","G","H","I","J"]:
            c(f"{col}9", 1)

        # Categorías de materiales (filas 10-20)
        categorias = [
            (10, "DESINFECCION Y FERTILIZACION"),
            (11, "AMPLIACION "),
            (12, "CULTIVO TIERRA, CHAROLAS"),
            (13, "MATERIAL VEGETAL"),
            (14, "PREPARACION DE SUELO"),
            (15, "FERTILIZANTES (Manejo Integrado de Riego y Fertilización) "),
            (16, "DESINFECCION / PLAGUICIDAS (Manejo Integrado de Plagas y Enfermedades)"),
            (17, "MANTENIMIENTO"),
            (18, "EXPANSION CECILIA 25"),
            (19, "RENOVACION DE SIEMBRA"),
            (20, "MATERIAL DE EMPAQUE"),
        ]
        for row, label in categorias:
            c(f"B{row}", label)

        c("B22", "COSTO DE MATERIALES")

        # Nóminas (filas 24-59)
        nominas = [
            (24, "NOMINA ADMON Oficina, Jefes de Finca, Ingenieros"),
            (25, "HORAS EXTR. DOM. Y FESTIVOS"),
            (26, "BONOS ASISIT, PUNTAULIDAD Y DESPENSA"),
            (27, "NOMINA PRODUCCION "),
            (28, "HORAS EXTR. DOM. Y FEST."),
            (29, "BONOS ASISIT, PUNT. Y DESP."),
            (30, "NOMINA PRODUCCION CORTE"),
            (31, "HORAS EXTR. DOM. Y FESTIVOS CORTE"),
            (32, "BONOS ASISIT, PUNTAULIDAD Y DESP. CORTE"),
            (33, "NOMINA PRODUCCION TRANSPLANTE"),
            (34, "HORAS EXTR. DOM. Y FEST. TRANSPLANTE"),
            (35, "BONOS ASISIT, PUNT. Y DESP. TRANSPLANTE"),
            (36, "NOMINA PRODUCCION MANEJO PLANTA"),
            (37, "HORAS EXTR. DOM. Y FEST. MANEJO PLANTA"),
            (38, "BONOS ASISIT, PUNT. Y DESP. MANEJO PLANTA"),
            (39, "NOMINA  HOOPS"),
            (40, "HORAS EXTR. DOM. Y FEST. HOOPS"),
            (41, "BONOS ASISIT, PUNT. Y DESP.HOOPS"),
            (42, "NOMINA  (MIPE,MIRFE,)"),
            (43, "HORAS EXTR. DOM. Y FEST. (MIPE,MIRFE)"),
            (44, "BONOS ASISIT, PUNT. Y DESP.(MIPE,MIRFE)"),
            (45, "NOMINA OPERATIVOS (TRACTORES, CAMEROS)"),
            (46, "HORAS EXTR. DOM. Y FEST. (TRACTORES, CAMEROS)"),
            (47, "BONOS ASISIT, PUNT. Y DESP. (TRACTORES, CAMEROS)"),
            (48, "NOMINA OPERATIVOS (CHOFER)"),
            (49, "HORAS EXTR. DOM. Y FEST. (CHOFER)"),
            (50, "BONOS ASISIT, PUNT. Y DESP. (CHOFER)"),
            (51, "NOMINA OPERATIVOS (VELADORES)"),
            (52, "HORAS EXTR. DOM. Y FEST. (VELADORES)"),
            (53, "BONOS ASISIT, PUNT. Y DESP. (VELADORES)"),
            (54, "NOMINA OPERATIVOS (SOLDADOR)"),
            (55, "HORAS EXTR. DOM. Y FEST. (SOLDADOR)"),
            (56, "BONOS ASISIT, PUNT. Y DESP. (SOLDADOR)"),
            (57, "NOMINA PRODUCCION Contratista y comisiones"),
            (58, "IMSS , INFONAVIT RCV"),
            (59, "1.8% al estado (1.2% tasa efectiva)"),
        ]
        for row, label in nominas:
            c(f"B{row}", label)

        c("B61", "COSTO DE MANO DE OBRA")

        # Servicios (filas 63-70)
        servicios = [
            (63, "ELECTRICIDAD"),
            (64, "FLETES Y ACARREOS (Flete aduana)"),
            (65, "GASTOS DE EXPORTACION "),
            (66, "CERTIFICADO DE FITOSANITARIOS"),
            (67, "Transporte de personal"),
            (68, "COMPRA DE FLOR A TERCEROS"),
            (69, "COMIDA PARA EL PERSONAL"),
            (70, "RO, TEL, RTA.ALIM."),
        ]
        for row, label in servicios:
            c(f"B{row}", label)

        c("B72", "COSTO DE SERVICIOS")
        c("B74", "COSTO DE PRODUCCION Y VENTAS")

        # Producción (filas 76-92)
        prod = [
            (76, "CAJAS PROCESADAS TOTALES"),
            (77, "INVENTARIO INICIAL"),
            (78, "TALLOS COSECHADOS"),
            (79, "TALLOS DESECHADOS"),
            (80, "TALLOS DESECHADOS sf"),
            (81, "TALLOS COMPRADOS"),
            (82, "TALLOS EN BOUQUETS O PROCESADOS"),
            (83, "TALLOS DESPACHADOS"),
            (84, "LIBRAS DESPACHADAS ALBAHACA"),
            (85, "TALLOS muestra"),
            (86, "INVENTARIO FINAL"),
            (87, "TALLOS PROCESADOS TOTALES"),
            (88, " CHAROLAS SEMBRADAS *288 PLUGS ="),
            (89, " NUMERO DE CHAROLAS SEMBRADAS "),
            (90, " NUMERO DE ESQUEJES SEMBRADOS"),
            (91, " METROS DE SIEMBRA"),
            (92, " HECTAREAS EN SIEMBRA"),
        ]
        for row, label in prod:
            c(f"B{row}", label)

        c("B93", "<<< INDICADORES")
        c("B94", "COSTOS UNITARIOS"); c("B95", "$ / Tallo Procesado")
        c("B96", "COSTOS UNITARIOS"); c("B97", "$ / Libras Procesadas")
        for row, label in [(98,"Materiales"),(99,"Mano de Obra"),(100,"Servicios (Fletes)"),
                           (101,"Costo de Produccion y Ventas"),(103,"Material de Empaque / Tallo"),
                           (105,"Sanidad Vegetal / Tallo"),(106,"Fertlizacion / Tallo"),
                           (108,"Mano de Obra Prod / Tallo")]:
            c(f"B{row}", label)
        c("B110", "$ / Hectarea")
        for row, label in [(111,"Materiales"),(112,"Mano de Obra"),(113,"Servicios (Fletes)"),
                           (114,"Costo de Produccion y Ventas"),(121,"Mano de Obra Prod / Ha")]:
            c(f"B{row}", label)

        c("B124", "KPI's ")
        c("B125", "Proyectos de inversión"); c("L125", "Total Weekly")
        proyectos = [
            (126,"Sistema de riego (Ramona)"),(127,"Sistema de riego (Isabella)"),
            (128,"Caseta (Isabella)"),(129,"Sistema de ventilacion)"),
            (130,"Sistema de tratamiento de aguas residuales (Isabella)"),
            (131,"Arcos para invernaderos "),(132,"proyecto luz"),
            (133,"Construcción de Almacén (Ramona) "),(134,"Construcción de Almacén (Isabela) "),
            (135,"Carritos"),(136,"Maquinaria "),(137,"Chiller"),
            (138,"Cuarto frio"),(139,"veronicas"),
        ]
        for row, label in proyectos:
            c(f"B{row}", label)
        c("B140", "Total ")
        c("B143", "Logística "); c("L143", "Total Weekly"); c("N143", "PosCo-RM")
        for row, label in [
            (144,"Número de camiones despachados "),(145,"Número de tarimas despachadas (montadas al camión)"),
            (146,"Número de cajas despachadas"),(147,"Número de Pies cúbicos de cajas despachadas "),
            (148,"Número de Pies cubicos promedio / camión despachado "),
            (149,"Capacidad en pies cúbicos por camión "),(150,"Rendimiento promedio por camión "),
        ]:
            c(f"B{row}", label)
        for row, label in [
            (152,"Costo incurrido por flete, gtos expo, fitosanitarios"),
            (153,"Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
            (154,"Número de Camiones despachados "),
            (156,"Costo incurrido promedio flete, gtos expo, fitosanitarios / pie cúbico"),
            (157,"Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
            (158,"Número de Pies cúbicos de cajas despachadas"),
            (160,"Costo incurrido flete, gtos expo, fitosanitarios / cajas despachadas"),
            (161,"Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
            (162,"Número de cajas despachadas"),
        ]:
            c(f"B{row}", label)
        c("B165", "Material de empaque / Caja")
        for row, label in [
            (166,"Costo incurrido en Material de empaque / pie cúbico"),
            (167,"Costo incurrido en Material de empaque (USD)"),
            (168,"Número de Pies cúbicos de cajas despachadas"),
            (170,"Costo incurrido en Material de empaque / cajas despachadas"),
            (171,"Costo incurrido en Material de empaque (USD)"),
            (172,"Número de cajas despachadas"),
        ]:
            c(f"B{row}", label)

        return celdas

    # ── 1. Token OAuth2 ───────────────────────────────────────────────────
    token_url = f'https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token'
    token_resp = requests.post(token_url, data={
        "grant_type":    "client_credentials",
        "client_id":     client_id,
        "client_secret": client_secret,
        "scope":         "https://graph.microsoft.com/.default",
    }, timeout=20)
    if token_resp.status_code != 200:
        return {"ok": False, "error": f"Error obteniendo token: {token_resp.text}"}

    token = token_resp.json().get('access_token')
    hdrs_json = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    # ── 2. Resolver driveId + itemId ──────────────────────────────────────
    encoded = _b64.b64encode(SHAREPOINT_URL_WK.encode()).decode().rstrip('=')
    encoded = 'u!' + encoded.replace('/', '_').replace('+', '-')
    res = requests.get(
        f'https://graph.microsoft.com/v1.0/shares/{encoded}/driveItem',
        headers=hdrs_json, timeout=20,
    )
    if res.status_code != 200:
        return {"ok": False, "error": f"No se pudo resolver el archivo: {res.text}"}

    item     = res.json()
    drive_id = item.get('parentReference', {}).get('driveId')
    item_id  = item.get('id')
    if not drive_id or not item_id:
        return {"ok": False, "error": "No se pudo obtener driveId o itemId."}

    wb_url = f'https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/workbook'

    # ── 3. Abrir sesión de workbook (permite trabajar con archivo abierto) ─
    sess_resp = requests.post(
        f'{wb_url}/createSession',
        headers=hdrs_json,
        json={"persistChanges": True},
        timeout=30,
    )
    if sess_resp.status_code not in (200, 201):
        return {"ok": False, "error": f"Error abriendo sesión: {sess_resp.text}"}

    session_id = sess_resp.json().get('id')
    hdrs = {**hdrs_json, "workbook-session-id": session_id}

    try:
        # ── 4. Verificar que la hoja no exista ────────────────────────────
        sheets_resp = requests.get(f'{wb_url}/worksheets', headers=hdrs, timeout=20)
        if sheets_resp.status_code != 200:
            return {"ok": False, "error": f"Error listando hojas: {sheets_resp.text}"}
        nombres = [h['name'].strip() for h in sheets_resp.json().get('value', [])]
        if nombre_hoja.upper() in [n.upper() for n in nombres]:
            return {"ok": False, "error": f"La hoja '{nombre_hoja}' ya existe."}

        # ── 5. Crear la hoja nueva ────────────────────────────────────────
        add_resp = requests.post(
            f'{wb_url}/worksheets/add',
            headers=hdrs,
            json={"name": nombre_hoja},
            timeout=20,
        )
        if add_resp.status_code not in (200, 201):
            return {"ok": False, "error": f"Error creando hoja: {add_resp.text}"}

        # ── 6. Mover la hoja al inicio (posición 0) ───────────────────────
        ws_id = add_resp.json().get('id', nombre_hoja)
        # Graph API usa el nombre como id en la URL
        move_resp = requests.patch(
            f'{wb_url}/worksheets/{nombre_hoja}',
            headers=hdrs,
            json={"position": 0},
            timeout=20,
        )
        # No es fatal si falla el reordenamiento

        # ── 7. Escribir celdas en lotes (batchUpdate vía range) ───────────
        #   Graph API permite escribir un rango completo de una vez con:
        #   PATCH /workbook/worksheets/{id}/range(address='A1:Z200')
        #   Body: { "values": [[row0col0, row0col1, ...], [row1col0, ...]] }
        #
        #   Construimos una matriz 175 x 19 (filas 1-175, cols A-S)
        NROWS, NCOLS = 175, 19  # cols A(0)..S(18)
        col_idx = {c: i for i, c in enumerate("ABCDEFGHIJKLMNOPQRS")}
        matrix = [[""] * NCOLS for _ in range(NROWS)]

        for cell in _celdas_de_la_hoja(nombre_hoja):
            addr = cell["address"]          # ej "B3"
            val  = cell["value"]
            # Parsear dirección
            col_str = ''.join(ch for ch in addr if ch.isalpha())
            row_str = ''.join(ch for ch in addr if ch.isdigit())
            if col_str in col_idx and row_str:
                r = int(row_str) - 1
                col_c = col_idx[col_str]
                if 0 <= r < NROWS and 0 <= col_c < NCOLS:
                    matrix[r][col_c] = val if val is not None else ""

        range_addr = f"A1:S{NROWS}"
        patch_resp = requests.patch(
            f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{range_addr}\')',
            headers=hdrs,
            json={"values": matrix},
            timeout=60,
        )
        if patch_resp.status_code not in (200, 201):
            return {"ok": False, "error": f"Error escribiendo celdas: {patch_resp.text}"}

        # ── 8. Aplicar formatos via Graph API ─────────────────────────────
        def fmt(rng, body):
            """PATCH de formato sobre un rango."""
            requests.patch(
                f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{rng}\')/format',
                headers=hdrs, json=body, timeout=30,
            )

        def fill(rng, color):
            """Color de fondo (hex sin #)."""
            requests.patch(
                f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{rng}\')/format/fill',
                headers=hdrs, json={"color": f"#{color}"}, timeout=30,
            )

        def font(rng, bold=False, color=None):
            body = {"bold": bold}
            if color:
                body["color"] = f"#{color}"
            requests.patch(
                f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{rng}\')/format/font',
                headers=hdrs, json=body, timeout=30,
            )

        # Ancho de columna B (muy ancha para etiquetas)
        requests.patch(
            f'{wb_url}/worksheets/{nombre_hoja}/columns/range(address=\'B:B\')/format',
            headers=hdrs, json={"columnWidth": 500}, timeout=30,
        )

        # ── Colores de fondo por sección ──────────────────────────────────
        # Azul claro — encabezado semana
        fill("B2",        "DAE3F3")
        # Verde lima — código semana
        fill("B3",        "C5E0B4")
        # Verde claro USD — columnas L:S en secciones de datos
        for rng in [
            "L5:R5",
            "L7:S7", "L8:S8",
            "L10:S20",
            "L22:S22",
            "L24:S59",
            "L61:S61",
            "L63:S70",
            "L72:S72",
            "L74:S74",
        ]:
            fill(rng, "CCFFCC")
        # Amarillo — sección producción USD
        for rng in ["L76:S92", "L88:S92"]:
            fill(rng, "FFFFCC")
        # Naranja — filas de subtotales
        for rng in ["C22:J22", "L22:S22",
                    "C61:J61", "L61:S61",
                    "C72:J72", "L72:S72"]:
            fill(rng, "FFCC99")

        # ── Negritas ──────────────────────────────────────────────────────
        # Fila 1, 2, 3 col B
        font("B1:B3", bold=True)
        font("B9",    bold=True)
        font("B22",   bold=True)
        font("B61",   bold=True)
        font("B72",   bold=True)
        font("B74",   bold=True)
        font("B93",   bold=True)
        font("B94",   bold=True)
        font("B95",   bold=True)
        font("B96",   bold=True)
        font("B97",   bold=True)
        font("B101",  bold=True)
        font("B110",  bold=True)
        font("B124",  bold=True)
        font("B140",  bold=True)
        # Totales de columna C en cada sección
        for rng in ["C22:J22", "L22:S22",
                    "C61:J61", "L61:S61",
                    "C72:J72", "L72:S72",
                    "C74:J74", "L74:S74"]:
            font(rng, bold=True)
        # Encabezados KPI verdes
        font("B125", bold=True, color="008000")
        font("L125", bold=True, color="008000")
        font("B143", bold=True, color="008000")
        font("L143", bold=True, color="008000")
        font("N143", bold=True, color="008000")
        font("B165", bold=True, color="008000")

        # ── Alineación centrada en encabezados ────────────────────────────
        fmt("B2",    {"horizontalAlignment": "Center"})
        fmt("B3",    {"horizontalAlignment": "Center"})
        fmt("C5:J5", {"horizontalAlignment": "Center"})
        fmt("L5:R5", {"horizontalAlignment": "Center"})
        fmt("B6",    {"horizontalAlignment": "Center"})
        fmt("C7:J7", {"horizontalAlignment": "Center"})
        fmt("L7:S7", {"horizontalAlignment": "Center"})
        fmt("C8",    {"horizontalAlignment": "Center"})
        fmt("L8",    {"horizontalAlignment": "Center"})
        fmt("B9",    {"horizontalAlignment": "Center"})
        fmt("L125",  {"horizontalAlignment": "Center"})
        fmt("L143",  {"horizontalAlignment": "Center"})
        fmt("N143",  {"horizontalAlignment": "Center"})

        # ── Merge de celdas ───────────────────────────────────────────────
        merges = [
            "C5:J5", "L5:R5",
            "C153:C154", "L153:L154",
            "C157:C158", "L157:L158",
            "C161:C162", "L161:L162",
            "C167:C168", "L167:L168",
            "C171:C172", "L171:L172",
        ]
        for m in merges:
            requests.post(
                f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{m}\')/merge',
                headers=hdrs, json={"across": False}, timeout=20,
            )

    finally:
        # ── 8. Cerrar la sesión siempre ───────────────────────────────────
        requests.post(
            f'{wb_url}/closeSession',
            headers=hdrs,
            timeout=20,
        )

    return {
        "ok": True,
        "mensaje": f"Hoja '{nombre_hoja}' creada exitosamente en SharePoint.",
    }


# ─── Descarga de una hoja WK#### como xlsx con formato completo ───────────────
def get_sheet_xlsx(week_code: str) -> bytes | None:
    """
    Descarga el Excel de SharePoint y extrae la hoja WK{week_code}
    como un archivo .xlsx independiente con formato completo.
    """
    archivo = _descargar_excel(SHAREPOINT_URL_WK, "Excel WK")
    if archivo is None:
        return None

    archivo_bytes = archivo.getvalue()
    sheet_name = f"WK{week_code}"

    try:
        wb = openpyxl.load_workbook(BytesIO(archivo_bytes))

        target = None
        for sname in wb.sheetnames:
            normalized = re.sub(r'\s+', '', sname.strip()).upper()
            if normalized == sheet_name.upper():
                target = sname
                break

        if target is None:
            return None

        src_ws = wb[target]
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = target

        for row in src_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font          = copy(cell.font)
                    new_cell.border        = copy(cell.border)
                    new_cell.fill          = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection    = copy(cell.protection)
                    new_cell.alignment     = copy(cell.alignment)

        for merge in src_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merge))

        for col_letter, col_dim in src_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width  = col_dim.width
            new_ws.column_dimensions[col_letter].hidden = col_dim.hidden
        for row_num, row_dim in src_ws.row_dimensions.items():
            new_ws.row_dimensions[row_num].height = row_dim.height
            new_ws.row_dimensions[row_num].hidden = row_dim.hidden

        buf = BytesIO()
        new_wb.save(buf)
        buf.seek(0)
        return buf.read()

    except Exception as e:
        print(f"⚠️  Error extrayendo hoja {sheet_name}: {e}")
        return None
