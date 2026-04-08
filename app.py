"""
data_extractor.py
Centro Floricultor de Baja California
- Hojas WK  → Excel en SharePoint/OneDrive (pandas + requests)
- Hojas PR  → Excel separado en SharePoint  (pandas + requests)
- Hojas MP  → Excel separado en SharePoint  (MANTENIMIENTO)
- Hojas ME  → Excel separado en SharePoint  (MATERIAL DE EMPAQUE)
"""

import re
import requests
import pandas as pd
import openpyxl
from copy import copy
from io import BytesIO

# ─── URLs de SharePoint ───────────────────────────────────────────────────────
# Archivo principal: hojas WK####
SHAREPOINT_URL_WK = (
    "https://pacificafarms-my.sharepoint.com/:x:/g/personal/"
    "anahi_mora_cfbc_co/IQAQCb79SzHtRrTQR71pSNQcASOWqFXyeGGzEhUcT9FRRJ4?e=ClxLCN"
)

# Archivo secundario: hojas PR####, MP####, ME####
SHAREPOINT_URL_PR = (
    "https://pacificafarms-my.sharepoint.com/:x:/g/personal/"
    "jesus_sandoval_cfbc_co/IQCecMwUnigFQa1m-0AYEw-rAenSSKPasiHLi1p2cqtPHkc?e=wpBfv7"
)

# Conteo de personal (Mano de Obra)
SHAREPOINT_URL_CONTEO = (
    "https://pacificafarms-my.sharepoint.com/:x:/g/personal/"
    "anahi_mora_cfbc_co/IQCZHoO8krj-R538RArePPMhAd-aSdBCsF2bPjd7clqUfbE?e=7P5ex4"
)

# ─── Constantes ───────────────────────────────────────────────────────────────
RANCH_CONFIG = {
    "Prop-RM":     {"color": "#047857", "codes": ["VIV"], "keywords": ["PROP"]},
    "PosCo-RM":    {"color": "#1d4ed8", "codes": ["POS", "LIM"], "keywords": ["POSCO"]},
    "Campo-RM":    {"color": "#b45309", "codes": ["CAM", "RAM"], "keywords": ["CAMPO"]},
    "Isabela":     {"color": "#7c3aed", "codes": ["ISA"], "keywords": ["ISABEL"]},
    "HOOPS":       {"color": "#c2410c", "codes": ["HOO"], "keywords": ["HOOPS"]},
    "Cecilia":     {"color": "#be185d", "codes": ["CEC"], "keywords": ["CECILIA"]},
    "Cecilia 25":  {"color": "#047857", "codes": ["C25"], "keywords": ["CECILIA 25"]},
    "Christina":   {"color": "#0369a1", "codes": ["CHR"], "keywords": ["CHRISTINA"]},
    "Albahaca-RM": {"color": "#6d28d9", "codes": ["ALB"], "keywords": ["ALBAHACA"]},
    "Campo-VI":    {"color": "#64748b", "codes": [], "keywords": ["CAMPO-VI", "CAMPO-IV"]}
}

RANCH_KEYS = []
for data in RANCH_CONFIG.values():
    RANCH_KEYS.extend(data["keywords"])

RANCH_CODE_MAP = {}
for ranch, data in RANCH_CONFIG.items():
    for code in data["codes"]:
        RANCH_CODE_MAP[code] = ranch

CATEGORIAS_ORDEN = [
    "DESINFECCION Y FERTILIZACION",
    "AMPLIACION",
    "CULTIVO TIERRA, CHAROLAS",
    "MATERIAL VEGETAL",
    "PREPARACION DE SUELO",
    "FERTILIZANTES",
    "DESINFECCION / PLAGUICIDAS",
    "MANTENIMIENTO",
    "EXPANSION CECILIA 25",
    "RENOVACION DE SIEMBRA",
    "MATERIAL DE EMPAQUE",
    "COSTO SERVICIOS",
    "COSTO MANO DE OBRA",
]

SKIP = {"ACUMULADO", "GRAFICOS I-IV", "COMPARATIVO", "DATOS", "HOJA1", "SHEET1"}


# ─── Descarga de Excel desde SharePoint ──────────────────────────────────────
def _descargar_excel(url: str, label: str = "archivo") -> BytesIO | None:
    """
    Descarga un archivo .xlsx desde un link público de SharePoint/OneDrive.
    Agrega el parámetro download=1 necesario para la descarga directa.
    Funciona con URLs que tengan o no el token ?e=...
    """
    url = url.strip()
    if "?e=" in url:
        download_url = url.replace("?e=", "?download=1&e=")
    elif "?" in url:
        download_url = url + "&download=1"
    else:
        download_url = url + "?download=1"
    try:
        response = requests.get(download_url, timeout=30)
        response.raise_for_status()
        # SharePoint a veces devuelve HTML si la URL no es válida
        content_type = response.headers.get("Content-Type", "")
        if "html" in content_type.lower():
            print(f"❌ {label}: SharePoint devolvió HTML en vez de Excel. Verifica que el link sea público.")
            return None
        return BytesIO(response.content)
    except Exception as e:
        print(f"❌ Error descargando {label}: {e}")
        return None


# Alias para compatibilidad con get_sheet_xlsx
def descargar_excel() -> BytesIO | None:
    return _descargar_excel(SHAREPOINT_URL_WK, "Excel WK")


def _leer_hoja(xls: pd.ExcelFile, titulo: str, rango_filas: int = 60,
               rango_cols: int = 35) -> list[list]:
    """
    Lee una hoja del ExcelFile y la retorna como lista de listas.
    Las celdas vacías / NaN se convierten a "".
    """
    try:
        df = pd.read_excel(
            xls,
            sheet_name=titulo,
            header=None,
            nrows=rango_filas,
        ).fillna("")
        if df.shape[1] > rango_cols:
            df = df.iloc[:, :rango_cols]
        return df.values.tolist()
    except Exception as e:
        print(f"   ⚠️  Error leyendo hoja '{titulo}': {e}")
        return []


# ─── Helpers de normalización ─────────────────────────────────────────────────
def norm_ranch(s: str):
    s = str(s).upper().strip()
    if "CAMPO-VI" in s or "CAMPO-IV" in s:               return "Campo-VI"
    if "CECILIA 25" in s:                                return "Cecilia 25"
    if "CECILIA" in s and "25" not in s:                 return "Cecilia"
    if "CAMPO" in s and "VI" not in s and "IV" not in s: return "Campo-RM"
    
    for ranch, data in RANCH_CONFIG.items():
        if ranch in ["Campo-VI", "Cecilia 25", "Cecilia", "Campo-RM"]: 
            continue
        for kw in data["keywords"]:
            if kw in s:
                return ranch
    return None


def norm_cat(s: str):
    s = str(s).upper().strip()
    if "DESINFECCION" in s and "FERTILIZ" in s:  return "DESINFECCION Y FERTILIZACION"
    if s.startswith("AMPLIACION"):                return "AMPLIACION"
    if "CULTIVO" in s:                            return "CULTIVO TIERRA, CHAROLAS"
    if "MATERIAL VEG" in s:                       return "MATERIAL VEGETAL"
    if "PREPARACION" in s:                        return "PREPARACION DE SUELO"
    if "FERTILIZANTE" in s:                       return "FERTILIZANTES"
    if "SANIDAD" in s or "PLAGUICIDA" in s:       return "DESINFECCION / PLAGUICIDAS"
    if "MANTENIMIENTO" in s:                      return "MANTENIMIENTO"
    if "EXPANSION" in s:                          return "EXPANSION CECILIA 25"
    if "RENOVACION" in s:                         return "RENOVACION DE SIEMBRA"
    if "MATERIAL DE EMP" in s:                    return "MATERIAL DE EMPAQUE"
    if "COSTO DE MAT" in s:                       return "COSTO_STOP"
    if "COSTO DE SERV" in s:                      return "COSTO SERVICIOS"
    if s.startswith("ELECTRICIDAD"):                        return "SV:Electricidad"
    if s.startswith("FLETES Y ACARREOS"):                   return "SV:Fletes y Acarreos"
    if s.startswith("GASTOS DE EXPORTACION"):               return "SV:Gastos de Exportación"
    if s.startswith("CERTIFICADO DE FITOSANITARIO"):        return "SV:Certificado Fitosanitario"
    if s.startswith("TRANSPORTE DE PERSONAL"):              return "SV:Transporte de Personal"
    if s.startswith("COMPRA DE FLOR"):                      return "SV:Compra de Flor a Terceros"
    if s.startswith("COMIDA PARA EL PERSONAL"):             return "SV:Comida para el Personal"
    if s.startswith("RO, TEL") or s.startswith("RO , TEL"): return "SV:RO, TEL, RTA.Alim"
    # ── MANO DE OBRA ──────────────────────────────────────────────────────────
    if "NOMINA" in s or "NÓMINA" in s:
        if "ADMON" in s:                               return "MO:Nómina Admon"
        if "CONTRATISTA" in s:                         return "MO:Nómina Prod. Contratista"
        if "CORTE" in s:                               return "MO:Nómina Prod. Corte"
        if "TRANSPLANTE" in s:                         return "MO:Nómina Prod. Transplante"
        if "MANEJO PLANTA" in s:                       return "MO:Nómina Prod. Manejo Planta"
        if "HOOPS" in s:                               return "MO:Nómina HOOPS"
        if "MIPE" in s or "MIRFE" in s:                return "MO:Nómina MIPE/MIRFE"
        if "TRACTORES" in s or "CAMEROS" in s:         return "MO:Nómina Op. Tractores/Cameros"
        if "CHOFER" in s:                              return "MO:Nómina Op. Chofer"
        if "VELADOR" in s:                             return "MO:Nómina Op. Veladores"
        if "SOLDADOR" in s:                            return "MO:Nómina Op. Soldador"
        if "PRODUCCION" in s or "PRODUCCIÓN" in s:     return "MO:Nómina Producción"
    if "HORAS EXTR" in s:
        if "CORTE" in s:                               return "MO:H.Extra Corte"
        if "TRANSPLANTE" in s:                         return "MO:H.Extra Transplante"
        if "MANEJO PLANTA" in s:                       return "MO:H.Extra Manejo Planta"
        if "HOOPS" in s:                               return "MO:H.Extra HOOPS"
        if "MIPE" in s or "MIRFE" in s:                return "MO:H.Extra MIPE/MIRFE"
        if "TRACTORES" in s or "CAMEROS" in s:         return "MO:H.Extra Tractores/Cameros"
        if "CHOFER" in s:                              return "MO:H.Extra Chofer"
        if "VELADOR" in s:                             return "MO:H.Extra Veladores"
        if "SOLDADOR" in s:                            return "MO:H.Extra Soldador"
        if "FESTIVOS" in s and "FEST." not in s:       return "MO:H.Extra Dom. y Festivos (Admon)"
        return                                                "MO:H.Extra Dom. y Fest. (Prod.)"
    if "BONOS ASISIT" in s:
        if "CORTE" in s:                               return "MO:Bonos Corte"
        if "TRANSPLANTE" in s:                         return "MO:Bonos Transplante"
        if "MANEJO PLANTA" in s:                       return "MO:Bonos Manejo Planta"
        if "HOOPS" in s:                               return "MO:Bonos HOOPS"
        if "MIPE" in s or "MIRFE" in s:                return "MO:Bonos MIPE/MIRFE"
        if "TRACTORES" in s or "CAMEROS" in s:         return "MO:Bonos Tractores/Cameros"
        if "CHOFER" in s:                              return "MO:Bonos Chofer"
        if "VELADOR" in s:                             return "MO:Bonos Veladores"
        if "SOLDADOR" in s:                            return "MO:Bonos Soldador"
        if "DESPENSA" in s:                            return "MO:Bonos Asist./Puntualidad (Admon)"
        return                                                "MO:Bonos Asist./Puntualidad (Prod.)"
    if "IMSS" in s or "INFONAVIT" in s:                return "MO:IMSS/INFONAVIT RCV"
    if "1.8%" in s or "TASA EFECTIVA" in s:            return "MO:1.8% Estado"
    return None


def sv(v) -> float:
    try:
        if isinstance(v, str):
            v = v.replace("$", "").replace(",", "").strip()
        f = float(v)
        return f if f == f else 0.0
    except (TypeError, ValueError):
        return 0.0


# ─── Parser genérico compartido (PR / MP / ME tienen el mismo formato) ────────
def _parse_generic(rows: list) -> dict:
    """
    Formato común a PR####, MP####, ME####:
      Col 2: UBICACION  (ej: RAMMIPRNN, CECMIPSNF)
      Col 5: PRODUCTO
      Col 7: UNIDADES
      Col 9: GASTO
    Retorna: { rancho: { tipo: [[producto, unidades, gasto, ubicacion], ...] } }
    """
    UBICACION_COL = 2
    PRODUCTO_COL  = 5
    UNIDADES_COL  = 7
    GASTO_COL     = 9

    result = {}
    accum  = {}   # (rancho, tipo, producto, ubicacion) → [u_total, g_total]

    for row in rows:
        if not row or len(row) < 10:
            continue

        ubicacion = str(row[UBICACION_COL]).strip().upper() if len(row) > UBICACION_COL else ''
        ubicacion = re.sub(r'\s+', '', ubicacion)

        if not ubicacion or len(ubicacion) < 6:
            continue
        if not re.match(r'^[A-Z0-9]+$', ubicacion):
            continue

        ranch_code = ubicacion[:3]
        rancho = RANCH_CODE_MAP.get(ranch_code)

        if not rancho and ubicacion.startswith('VIV'):
            rancho = 'Prop-RM'

        if not rancho:
            continue

        tipo = 'MIPE' if 'MIP' in ubicacion else 'MIRFE'

        producto = str(row[PRODUCTO_COL]).strip() if len(row) > PRODUCTO_COL else ''
        if not producto or producto.upper() in ('PRODUCTO', 'NOMBRE', ''):
            continue

        unidades = str(row[UNIDADES_COL]).strip() if len(row) > UNIDADES_COL else ''
        try:
            u = float(str(unidades).replace(',', ''))
            unidades = str(int(u)) if u == int(u) else str(round(u, 2))
        except Exception:
            unidades = '0'

        gasto = str(row[GASTO_COL]).strip() if len(row) > GASTO_COL else ''
        try:
            g = float(str(gasto).replace(',', ''))
            gasto = str(round(g, 2))
        except Exception:
            gasto = '0'

        u_f = float(unidades) if unidades else 0.0
        g_f = float(gasto)    if gasto    else 0.0

        key = (rancho, tipo, producto, ubicacion)
        if key in accum:
            accum[key][0] += u_f
            accum[key][1] += g_f
        else:
            accum[key] = [u_f, g_f]

    for (rancho, tipo, producto, ubicacion), (u_tot, g_tot) in accum.items():
        u_str = str(int(u_tot)) if u_tot == int(u_tot) else str(round(u_tot, 2))
        g_str = str(round(g_tot, 2))
        result.setdefault(rancho, {}).setdefault(tipo, []).append([producto, u_str, g_str, ubicacion])

    return result


# ─── Fetch hojas PR / MP / ME desde el segundo Excel de SharePoint ────────────
def _fetch_desde_sharepoint(prefix: str, parser_fn, label: str) -> tuple[dict, dict]:
    """
    Descarga el Excel secundario de SharePoint y extrae todas las hojas
    que coincidan con el patrón  {PREFIX}####  (ej: PR2611, MP2608, ME2610).

    Args:
        prefix:    "PR", "MP" o "ME"
        parser_fn: función que convierte list[list] → dict de ranchos
        label:     nombre legible para logs

    Returns:
        (datos, debug)  con el mismo formato que antes usaban las funciones gspread
    """
    datos = {}
    debug = {f"hojas_{prefix.lower()}_encontradas": []}

    archivo = _descargar_excel(SHAREPOINT_URL_PR, f"Excel {label}")
    if archivo is None:
        print(f"⚠️  No se pudo descargar el archivo para hojas {prefix}")
        return datos, debug

    try:
        xls = pd.ExcelFile(archivo)
    except Exception as e:
        print(f"⚠️  No se pudo abrir el Excel de {label}: {e}")
        return datos, debug

    hojas_encontradas = []
    pat = re.compile(rf'^{prefix}\s*\d{{4}}$', re.IGNORECASE)

    for sname in xls.sheet_names:
        sname = sname.strip()
        if pat.match(sname):
            raw_code = re.sub(rf'{prefix}\s*', '', sname, flags=re.IGNORECASE).strip()
            try:
                code = int(raw_code)
                year = 2000 + (code // 100)
                if 2018 <= year <= 2030:
                    print(f"   ✅ {prefix}{code} encontrada en SharePoint: {sname}")
                    hojas_encontradas.append((sname, code))
                else:
                    print(f"   ❌ {prefix}{code} año {year} fuera de rango")
            except ValueError as e:
                print(f"   ❌ Error código '{raw_code}': {e}")

    debug[f"hojas_{prefix.lower()}_encontradas"] = [t for t, _ in hojas_encontradas]

    if not hojas_encontradas:
        print(f"   ℹ️  No hay hojas {prefix} en el Excel de SharePoint")
        return datos, debug

    for titulo, code in hojas_encontradas:
        vals   = _leer_hoja(xls, titulo, rango_filas=500, rango_cols=11)
        parsed = parser_fn(vals)
        datos[code] = parsed
        debug[f"{prefix}{code}_ranchos"] = list(parsed.keys()) if parsed else []
        print(f"   📦 {prefix}{code} ranchos detectados: {list(parsed.keys())}")

    return datos, debug


# ─── Extractor principal ──────────────────────────────────────────────────────
def extraer_datos(xls: pd.ExcelFile) -> dict:
    all_data       = []
    servicios_data = []
    mano_obra_data = []

    hojas_validas = []
    pr_hojas      = []

    print("\n" + "=" * 60)
    print("🔍 DETECTANDO HOJAS EN EL EXCEL WK")
    print("=" * 60)

    for sname in xls.sheet_names:
        sname = sname.strip()
        print(f"\n📄 Hoja: '{sname}'")

        if sname.upper() in SKIP:
            print("   ⏭️  SKIP (en lista de exclusión)")
            continue

        pr_match = re.match(r'^PR\s*\d{4}$', sname, re.IGNORECASE)
        if pr_match:
            pr_raw = re.sub(r'PR\s*', '', sname, flags=re.IGNORECASE).strip()
            try:
                pr_code = int(pr_raw)
                pr_year = 2000 + (pr_code // 100)
                if 2018 <= pr_year <= 2030:
                    print("   ✅ PR DETECTADA Y VÁLIDA (en WK Excel)")
                    pr_hojas.append((sname, pr_code))
                    continue
            except ValueError:
                pass

        wk_match = re.match(r'^WK\s*\d{4}$', sname, re.IGNORECASE)
        if wk_match:
            code_raw = re.sub(r"WK\s*", "", sname, flags=re.IGNORECASE).strip()
            try:
                code = int(code_raw)
                year = 2000 + (code // 100)
                if 2018 <= year <= 2030:
                    print("   ✅ WK DETECTADA Y VÁLIDA")
                    hojas_validas.append((sname, code))
                else:
                    print(f"   ❌ Año {year} fuera de rango")
            except ValueError:
                print("   ❌ Error convirtiendo código")
        else:
            if not pr_match:
                print("   ℹ️  No es WK ni PR")

    print("\n" + "=" * 60)
    print("📊 RESUMEN:")
    print(f"   • Hojas WK encontradas: {len(hojas_validas)}")
    print(f"   • Hojas PR en WK Excel: {len(pr_hojas)}")
    print("=" * 60 + "\n")

    if not hojas_validas:
        return {"error": "No se encontraron hojas WK validas."}

    # 2. Leer hojas WK
    batch_data = {}
    for titulo, _ in hojas_validas:
        batch_data[titulo] = _leer_hoja(xls, titulo, rango_filas=120, rango_cols=35)

    # 2b. Leer hojas PR que estén en el Excel WK (fallback)
    productos       = {}
    productos_debug = {"hojas_pr_encontradas": [t for t, _ in pr_hojas]}
    for titulo, pr_code in pr_hojas:
        vals   = _leer_hoja(xls, titulo, rango_filas=500, rango_cols=11)
        parsed = _parse_generic(vals)
        productos[pr_code] = parsed
        productos_debug[f"PR{pr_code}_ranchos"] = list(parsed.keys()) if parsed else []

    # 3. Procesar cada hoja WK
    siembra_data: dict = {}  # {wk_code: {ranch: {charolas,esquejes,metros,hectareas}}}
    SIEMBRA_LABELS = [
        ("inv_inicial",     "INVENTARIO INICIAL"),
        ("tallos_cos",      "TALLOS COSECHADOS"),
        ("tallos_des",      "TALLOS DESECHADOS"),
        ("tallos_des_sf",   "TALLOS DESECHADOS SF"),
        ("tallos_comp",     "TALLOS COMPRADOS"),
        ("tallos_bouq",     "TALLOS EN BOUQUETS O PROCESADOS"),
        ("tallos_desp",     "TALLOS DESPACHADOS"),
        ("libras_alb",      "LIBRAS DESPACHADAS ALBAHACA"),
        ("tallos_mues",     "TALLOS MUESTRA"),
        ("inv_final",       "INVENTARIO FINAL"),
        ("tallos_proc",     "TALLOS PROCESADOS TOTALES"),
        ("charolas_288",    "CHAROLAS SEMBRADAS"),
        ("charolas",        "NUMERO DE CHAROLAS SEMBRADAS"),
        ("esquejes",        "NUMERO DE ESQUEJES SEMBRADOS"),
        ("metros",          "METROS DE SIEMBRA"),
        ("hectareas",       "HECTAREAS EN SIEMBRA"),
    ]

    for titulo, code in hojas_validas:
        raw = batch_data.get(titulo, [])
        if not raw:
            continue

        yy   = code // 100
        ww   = code % 100
        year = 2000 + yy

        max_cols = max((len(r) for r in raw), default=0)
        data     = [r + [""] * (max_cols - len(r)) for r in raw]

        date_range = ""
        for _dr in range(min(8, len(data))):
            for _dc in range(min(5, len(data[_dr]))):
                _v = str(data[_dr][_dc]).strip()
                if _v and " al " in _v.lower() and len(_v) > 8:
                    date_range = _v
                    break
            if date_range:
                break

        exec_idx = -1
        for i, row in enumerate(data):
            if any(isinstance(c, str) and "EJECUCION SEMANAL" in c.upper() for c in row):
                exec_idx = i
                break
        if exec_idx < 0:
            continue

        header_idx = -1
        for i in range(exec_idx - 1, max(0, exec_idx - 6) - 1, -1):
            if any(isinstance(v, str) and any(k in v.upper() for k in RANCH_KEYS) for v in data[i]):
                header_idx = i
                break
        if header_idx < 0:
            continue

        header = data[header_idx]

        total_cols = [j for j, v in enumerate(header)
                      if isinstance(v, str) and v.strip().upper() == "TOTAL"]
        if not total_cols:
            continue
        mxn_total_col = total_cols[0]
        usd_total_col = total_cols[1] if len(total_cols) >= 2 else None

        mxn_ranch_cols, usd_ranch_cols = {}, {}
        for j, v in enumerate(header):
            rn = norm_ranch(str(v)) if v else None
            if not rn:
                continue
            if j < mxn_total_col:
                mxn_ranch_cols[j] = rn
            elif usd_total_col and mxn_total_col < j < usd_total_col:
                mxn_ranch_cols[j] = rn
            elif usd_total_col and j > usd_total_col:
                usd_ranch_cols[j] = rn

        print(f"\n[DEBUG {titulo}]")
        print(f"   exec_idx={exec_idx}, header_idx={header_idx}")
        print(f"   mxn_total_col={mxn_total_col}, usd_total_col={usd_total_col}")
        print(f"   mxn_ranch_cols={mxn_ranch_cols}")
        print(f"   usd_ranch_cols={usd_ranch_cols}")
        hdr_vals = [(j, str(header[j])[:15]) for j in range(len(header)) if str(header[j]).strip()]
        print(f"   header non-empty: {hdr_vals}")

        for i in range(exec_idx + 1, min(exec_idx + 120, len(data))):
            row   = data[i]
            label = next((str(row[c]).strip() for c in range(5)
                          if c < len(row) and row[c] and len(str(row[c]).strip()) > 3), None)
            if not label:
                continue

            cat = norm_cat(label)
            if not cat:
                continue
            if cat == "COSTO_STOP":
                continue

            mxn_ranches = {rn: sv(row[j]) for j, rn in mxn_ranch_cols.items() if j < len(row)}
            usd_ranches = {rn: sv(row[j]) for j, rn in usd_ranch_cols.items() if j < len(row)}

            if cat.startswith("SV:"):
                print(f"   [SV] fila={i} label='{label[:30]}' cat='{cat}' mxn_ranches={mxn_ranches}")
                servicios_data.append({
                    "semana":      code,
                    "year":        year,
                    "week":        ww,
                    "date_range":  date_range,
                    "subcat":      cat[3:],
                    "mxn_total":   round(sv(row[mxn_total_col]) if mxn_total_col < len(row) else 0, 2),
                    "usd_total":   round(sv(row[usd_total_col]) if usd_total_col and usd_total_col < len(row) else 0, 2),
                    "mxn_ranches": mxn_ranches,
                    "usd_ranches": usd_ranches,
                })
            elif cat.startswith("MO:"):
                mano_obra_data.append({
                    "semana":      code,
                    "year":        year,
                    "week":        ww,
                    "date_range":  date_range,
                    "subcat":      cat[3:],
                    "mxn_total":   round(sv(row[mxn_total_col]) if mxn_total_col < len(row) else 0, 2),
                    "usd_total":   round(sv(row[usd_total_col]) if usd_total_col and usd_total_col < len(row) else 0, 2),
                    "mxn_ranches": mxn_ranches,
                    "usd_ranches": usd_ranches,
                })
            else:
                all_data.append({
                    "semana":      code,
                    "year":        year,
                    "week":        ww,
                    "date_range":  date_range,
                    "categoria":   cat,
                    "mxn_total":   round(sv(row[mxn_total_col]) if mxn_total_col < len(row) else 0, 2),
                    "usd_total":   round(sv(row[usd_total_col]) if usd_total_col and usd_total_col < len(row) else 0, 2),
                    "mxn_ranches": mxn_ranches,
                    "usd_ranches": usd_ranches,
                })

        # ── Extraer filas de siembra (89-92 Excel = labels fijos) ────────────
        wk_siembra: dict = {}
        for field_key, field_label in SIEMBRA_LABELS:
            for row in data:
                cell_text = " ".join(str(row[c]).strip().upper() for c in range(min(5, len(row))))
                if field_label in cell_text:
                    # Total MXN (columna total)
                    total_val = sv(row[mxn_total_col]) if mxn_total_col < len(row) else 0
                    wk_siembra.setdefault("TOTAL", {})[field_key] = total_val
                    # Por rancho
                    for j, rn in mxn_ranch_cols.items():
                        if j < len(row):
                            wk_siembra.setdefault(rn, {})[field_key] = sv(row[j])
                    break
        if wk_siembra:
            siembra_data[code] = wk_siembra

    print(f"\n✅ servicios_data: {len(servicios_data)} registros encontrados")
    if servicios_data:
        print(f"   subcats: {list({r['subcat'] for r in servicios_data})}")
    print(f"✅ mano_obra_data: {len(mano_obra_data)} registros encontrados")

    cats_found = {r["categoria"] for r in all_data}
    if mano_obra_data:
        cats_found.add("COSTO MANO DE OBRA")
    cats  = [c for c in CATEGORIAS_ORDEN if c in cats_found]
    years = sorted({r["year"] for r in all_data})

    ranches_seen: set = set()
    for r in all_data:
        ranches_seen.update(r["mxn_ranches"])
        ranches_seen.update(r["usd_ranches"])
    ranches = sorted(ranches_seen)

    summary: dict = {cat: {yr: {"usd": 0.0, "mxn": 0.0, "ranches": {}, "ranches_mxn": {}}
                            for yr in years} for cat in cats}
    for r in all_data:
        s = summary.get(r["categoria"], {}).get(r["year"])
        if not s:
            continue
        s["usd"] += r["usd_total"]
        s["mxn"] += r["mxn_total"]
        for rn, v in r["usd_ranches"].items():
            s["ranches"][rn] = round(s["ranches"].get(rn, 0) + v, 2)
        for rn, v in r["mxn_ranches"].items():
            s["ranches_mxn"][rn] = round(s["ranches_mxn"].get(rn, 0) + v, 2)
    for cat in cats:
        for yr in years:
            d = summary[cat][yr]
            d["usd"] = round(d["usd"], 2)
            d["mxn"] = round(d["mxn"], 2)

    weeks_per_year: dict = {}
    week_date_ranges: dict = {}
    for r in all_data:
        weeks_per_year.setdefault(r["year"], set()).add(r["week"])
        key = f"{r['year']}-{r['week']}"
        if key not in week_date_ranges and r.get("date_range"):
            week_date_ranges[key] = r["date_range"]
    weeks_per_year = {yr: sorted(wks) for yr, wks in weeks_per_year.items()}

    return {
        "years":            years,
        "categories":       cats,
        "ranches":          ranches,
        "summary":          summary,
        "weeks_per_year":   weeks_per_year,
        "week_date_ranges": week_date_ranges,
        "weekly_detail":    all_data,
        "productos":        productos,
        "productos_debug":  productos_debug,
        "servicios_data":   servicios_data,
        "mano_obra_data":   mano_obra_data,
        "siembra_data":     siembra_data,
    }


# ─── Punto de entrada público ─────────────────────────────────────────────────
def _extraer_mano_obra_conteo() -> list:
    """
    Lee el Excel de conteo de personal desde SharePoint.
    """
    print(f"📥 Descargando conteo desde: {SHAREPOINT_URL_CONTEO}")
    archivo = _descargar_excel(SHAREPOINT_URL_CONTEO, "Conteo Personal")
    if archivo is None:
        print("⚠️  No se pudo descargar conteo.xlsx — mano_obra_data vacío")
        return []
    print(f"✅ Descarga OK, tamaño={len(archivo.getvalue())} bytes")

    try:
        df = pd.read_excel(archivo, sheet_name="BD", header=2)
        print(f"✅ Excel leído: {df.shape[0]} filas, columnas={list(df.columns)}")
    except Exception as e:
        print(f"⚠️  Error leyendo conteo.xlsx: {e}")
        return []

    df.columns = [str(c).strip() for c in df.columns]
    needed = {"Año", "Semana", "Área", "Rancho", "Costo MN", "Costo DLLS", "Conteo"}
    missing = needed - set(df.columns)
    if missing:
        print(f"⚠️  Conteo.xlsx — columnas faltantes: {missing}")
        return []

    # ── Diagnóstico columna Conteo ────────────────────────────────────────────
    print(f"🔍 Conteo dtype  : {df['Conteo'].dtype}")
    print(f"🔍 Conteo sample : {df['Conteo'].head(10).tolist()}")
    print(f"🔍 Conteo no-nulos: {df['Conteo'].notna().sum()} de {len(df)} filas")
    # Forzar conversión numérica por si llegan como string o formula
    df["Conteo"] = pd.to_numeric(df["Conteo"], errors="coerce").fillna(0.0)

    df = df.dropna(subset=["Año", "Semana", "Área"])
    df["Año"]    = pd.to_numeric(df["Año"],    errors="coerce")
    df["Semana"] = pd.to_numeric(df["Semana"], errors="coerce")
    df = df.dropna(subset=["Año", "Semana"])
    df["Año"]    = df["Año"].astype(int)
    df["Semana"] = df["Semana"].astype(int)

    def _sv(v):
        try:
            s = str(v).strip().replace(",", "").replace(" ", "")
            if not s or s in ("-", "-   ", " -   "):
                return 0.0
            return round(float(s), 2)
        except:
            return 0.0

    result = []
    for (anio, semana, area), grp in df.groupby(["Año", "Semana", "Área"]):
        code = (int(anio) - 2000) * 100 + int(semana)
        mxn_ranches, usd_ranches, hc_ranches = {}, {}, {}
        mxn_total = usd_total = hc_total = 0.0
        for _, row in grp.iterrows():
            rancho     = str(row.get("Rancho", "")).strip()
            costo_mn   = _sv(row.get("Costo MN",   0))
            costo_dlls = _sv(row.get("Costo DLLS", 0))
            conteo_val = _sv(row.get("Conteo", 0))
            mxn_total += costo_mn
            usd_total += costo_dlls
            hc_total  += conteo_val
            if rancho:
                mxn_ranches[rancho] = round(mxn_ranches.get(rancho, 0.0) + costo_mn,   2)
                usd_ranches[rancho] = round(usd_ranches.get(rancho, 0.0) + costo_dlls, 2)
                hc_ranches[rancho]  = hc_ranches.get(rancho, 0.0) + conteo_val
        result.append({
            "semana":      code,
            "year":        int(anio),
            "week":        int(semana),
            "date_range":  "",
            "subcat":      str(area).strip(),
            "mxn_total":   round(mxn_total, 2),
            "usd_total":   round(usd_total, 2),
            "hc_total":    hc_total,
            "mxn_ranches": mxn_ranches,
            "usd_ranches": usd_ranches,
            "hc_ranches":  hc_ranches,
        })

    print(f"✅ conteo mano_obra_data: {len(result)} registros desde conteo.xlsx")
    return result


def get_datos() -> dict:
    """
    - Hojas WK  → Excel principal en SharePoint
    - Hojas PR  → Excel secundario en SharePoint
    - Hojas MP  → Excel secundario en SharePoint (MANTENIMIENTO)
    - Hojas ME  → Excel secundario en SharePoint (MATERIAL DE EMPAQUE)
    """
    # 1. Descargar y leer Excel WK
    archivo = _descargar_excel(SHAREPOINT_URL_WK, "Excel WK")
    if archivo is None:
        return {"error": "No se pudo descargar el archivo WK de SharePoint."}

    try:
        xls = pd.ExcelFile(archivo)
    except Exception as e:
        return {"error": f"No se pudo abrir el Excel WK: {e}"}

    resultado = extraer_datos(xls)

    if "error" not in resultado:
        # 2. Leer PR desde Excel secundario de SharePoint
        print("\n" + "=" * 60)
        print("🔍 LEYENDO HOJAS PR DESDE SHAREPOINT")
        print("=" * 60)
        productos, productos_debug = _fetch_desde_sharepoint("PR", _parse_generic, "PR")
        # Merge con cualquier PR que ya estuviera en el Excel WK
        resultado["productos"].update(productos)
        resultado["productos_debug"].update(productos_debug)

        # 3. Leer MP desde Excel secundario de SharePoint (MANTENIMIENTO)
        print("\n" + "=" * 60)
        print("🔍 LEYENDO HOJAS MP DESDE SHAREPOINT (MANTENIMIENTO)")
        print("=" * 60)
        productos_mp, productos_mp_debug = _fetch_desde_sharepoint("MP", _parse_generic, "MP")
        resultado["productos_mp"]       = productos_mp
        resultado["productos_mp_debug"] = productos_mp_debug

        # 4. Leer ME desde Excel secundario de SharePoint (MATERIAL DE EMPAQUE)
        print("\n" + "=" * 60)
        print("🔍 LEYENDO HOJAS ME DESDE SHAREPOINT (MATERIAL DE EMPAQUE)")
        print("=" * 60)
        productos_me, productos_me_debug = _fetch_desde_sharepoint("ME", _parse_generic, "ME")
        resultado["productos_me"]       = productos_me
        resultado["productos_me_debug"] = productos_me_debug

        HIDDEN_RANCHES = {"Albahaca-RM", "Campo-VI"}
        resultado["config"] = {
            "ranch_order": [k for k in RANCH_CONFIG.keys() if k not in HIDDEN_RANCHES],
            "ranch_colors": {k: v["color"] for k, v in RANCH_CONFIG.items() if k not in HIDDEN_RANCHES}
        }

        # 5. Reemplazar mano_obra_data con datos del conteo de personal
        print("\n" + "=" * 60)
        print("🔍 LEYENDO CONTEO DE PERSONAL (MANO DE OBRA)")
        print("=" * 60)
        resultado["mano_obra_data"] = _extraer_mano_obra_conteo()

    return resultado


# --- Construir hoja WK en blanco con estructura fija ---
def _construir_hoja_wk(ws, nombre_hoja: str):
    """
    Escribe la estructura completa de una hoja WK#### con formato IDENTICO al Excel de SharePoint.
    Colores, negritas, bordes y rellenos exactos.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # ── Helpers ─────────────────────────────────────────────────────────
    NAVY  = "333399"   # color texto principal
    GRAY  = "44546A"   # color texto datos rancho
    WHITE = "FFFFFF"

    def _f(bold=False, size=10, color=NAVY, name="Calibri"):
        return Font(bold=bold, size=size, color=color, name=name)

    def _fill(hex_color):
        if not hex_color or hex_color in ("", "none"):
            return PatternFill(fill_type=None)
        c = hex_color.lstrip("FF") if len(hex_color) == 8 else hex_color
        if len(c) != 6:
            return PatternFill(fill_type=None)
        return PatternFill("solid", fgColor=c)

    def _al(h="general", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    thin   = Side(style="thin")
    medium = Side(style="medium")
    none_s = Side(style=None)

    def _bdr(left=None, right=None, top=None, bottom=None):
        return Border(left=left or none_s, right=right or none_s,
                      top=top or none_s,   bottom=bottom or none_s)

    # Rellenos
    fill_green  = _fill("CCFFCC")   # verde claro USD   (FFCCFFCC en real)
    fill_blue   = _fill("DAE3F3")   # azul claro encabezado
    fill_lime   = _fill("C5E0B4")   # verde lima codigo semana
    fill_orange = _fill("FFCC99")   # naranja subtotales (FFFFCC99)
    fill_yellow = _fill("FFFFCC")   # amarillo produccion
    fill_white  = _fill("FFFFFF")
    fill_kpi    = _fill("008000")   # verde oscuro KPI headers

    # Bordes reutilizables
    bdr_L_med         = _bdr(left=medium)
    bdr_L_med_R_thin  = _bdr(left=medium, right=thin)
    bdr_R_med         = _bdr(right=medium)
    bdr_L_R_thin      = _bdr(left=thin, right=thin)

    # ── Ancho de columnas ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 69.4
    ws.column_dimensions["C"].width = 14
    for col in ("D","E","F","G","H","I","J"):
        ws.column_dimensions[col].width = 11
    ws.column_dimensions["K"].width = 3
    for col in ("L","M","N","O","P","Q","R","S"):
        ws.column_dimensions[col].width = 11

    # ── Fila 1 ───────────────────────────────────────────────────────────
    ws["B1"].value = "CENTRO FLORICULTOR DE BAJA CALIFORNIA, S.A. DE C.V. "
    ws["B1"].font  = _f(bold=True)

    # ── Fila 2 ───────────────────────────────────────────────────────────
    ws["B2"].value = "SEMANA DE CALCULO - Mexico"
    ws["B2"].font  = _f(bold=True)
    ws["B2"].fill  = fill_blue
    ws["B2"].alignment = _al("center")

    # ── Fila 3 ───────────────────────────────────────────────────────────
    code = nombre_hoja[2:] if nombre_hoja.upper().startswith("WK") else nombre_hoja
    ws["B3"].value = code
    ws["B3"].font  = _f(bold=True)
    ws["B3"].fill  = fill_lime
    ws["B3"].alignment = _al("center")
    ws["C3"].value = 19;  ws["C3"].font = _f(bold=True)
    ws["C3"].border = _bdr(bottom=medium)
    ws["D3"].value = " tipo de cambio"; ws["D3"].font = _f(bold=True)
    ws["L3"].value = 19;  ws["L3"].font = _f(bold=True)
    ws["M3"].value = "  tipo de cambio "; ws["M3"].font = _f(bold=True)

    # ── Fila 4 ───────────────────────────────────────────────────────────
    ws["B4"].value = "Del ___ al ___ de ________ 20__"
    ws["B4"].alignment = _al("center")
    ws.row_dimensions[4].height = 15

    # ── Fila 5 ───────────────────────────────────────────────────────────
    ws.merge_cells("C5:J5")
    ws["C5"].value = "(MXN) Pesos Mexicanos"
    ws["C5"].alignment = _al("center")
    ws["C5"].border = _bdr(left=medium, right=medium, top=medium, bottom=thin)
    ws.merge_cells("L5:R5")
    ws["L5"].value = "US Dollars"
    ws["L5"].fill  = fill_green
    ws["L5"].alignment = _al("center")
    ws["L5"].border = _bdr(left=medium, top=medium, bottom=thin)
    ws["S5"].fill  = fill_green
    ws["S5"].border = _bdr(right=medium, top=medium, bottom=thin)

    # ── Fila 6 ───────────────────────────────────────────────────────────
    ws["B6"].value = "TOTAL FINCA"
    ws["B6"].fill  = fill_white
    ws["B6"].alignment = _al("center")
    ws["B6"].border = bdr_L_med
    ws["C6"].border = _bdr(left=medium, right=thin, top=thin)
    for col in ("L","M","N","O","P","Q","R"):
        ws[f"{col}6"].fill = fill_green
        ws[f"{col}6"].alignment = _al("center")
    ws["L6"].border = _bdr(left=medium, right=thin)
    ws["S6"].fill  = fill_green
    ws["S6"].border = bdr_R_med
    ws["S6"].alignment = _al("center")
    ws.row_dimensions[6].height = 26.4

    # ── Fila 7 ───────────────────────────────────────────────────────────
    ws["B7"].value = "Produccion"
    ws["B7"].fill  = fill_white
    ws["B7"].alignment = _al("center")
    ws["B7"].border = bdr_L_med
    headers_mxn = ["TOTAL","Prop-RM","PosCo-RM","Campo -RM","Isabela","Christina","Cecilia","Cecilia 25"]
    headers_usd = ["TOTAL","Prop-RM","PosCo-RM","Campo -RM","ISABELA","Christina","CECILIA","CECILIA 25"]
    for i, h in enumerate(headers_mxn):
        col = chr(ord("C")+i)
        ws[f"{col}7"].value = h
        ws[f"{col}7"].font  = _f(bold=(i==0))
        ws[f"{col}7"].alignment = _al("center")
    ws["C7"].border = _bdr(left=medium, right=thin)
    ws["J7"].border = bdr_R_med
    usd_cols = ["L","M","N","O","P","Q","R","S"]
    for i, h in enumerate(headers_usd):
        c = usd_cols[i]
        ws[f"{c}7"].value = h
        ws[f"{c}7"].fill  = fill_green
        ws[f"{c}7"].alignment = _al("center")
    ws["L7"].border = _bdr(left=medium, right=thin)
    ws["S7"].border = bdr_R_med

    # ── Fila 8 ───────────────────────────────────────────────────────────
    ws["C8"].value = "SEMANAL "
    ws["C8"].alignment = _al("center")
    ws["C8"].border = _bdr(left=medium, right=thin)
    ws["L8"].value = '"WEEKLY"'
    ws["L8"].fill  = fill_green
    ws["L8"].alignment = _al("center")
    ws["L8"].border = _bdr(left=medium, right=thin)
    for col in ("M","N","O","P","Q","R"):
        ws[f"{col}8"].fill = fill_green
    ws["S8"].fill  = fill_green
    ws["S8"].border = bdr_R_med

    # ── Fila 9 ───────────────────────────────────────────────────────────
    ws["B9"].value = "EJECUCION SEMANAL"
    ws["B9"].font  = _f(bold=True)
    ws["B9"].fill  = fill_white
    ws["B9"].alignment = _al("center")
    ws["B9"].border = _bdr(left=medium, bottom=thin)
    ws["C9"].border = _bdr(left=medium, right=thin, bottom=thin)
    for col in ("D","E","F","G","H","I"):
        ws[f"{col}9"].value = 1
        ws[f"{col}9"].alignment = _al("center")
        ws[f"{col}9"].border = _bdr(bottom=thin)
    ws["J9"].value = 1
    ws["J9"].alignment = _al("center")
    ws["J9"].border = _bdr(right=medium, bottom=thin)
    ws["L9"].fill  = fill_green
    ws["L9"].border = _bdr(left=medium, right=thin, bottom=thin)
    for col in ("M","N","O","P","Q","R"):
        ws[f"{col}9"].fill  = fill_green
        ws[f"{col}9"].border = _bdr(bottom=thin)
    ws["S9"].fill  = fill_green
    ws["S9"].border = _bdr(right=medium, bottom=thin)

    # ── Helper fila de categoría ─────────────────────────────────────────
    def _fila_cat(row, label, fill_usd=None, top_border=False):
        if fill_usd is None:
            fill_usd = fill_green
        ws[f"B{row}"].value = label
        ws[f"B{row}"].font  = _f()
        ws[f"B{row}"].fill  = fill_white
        ws[f"B{row}"].alignment = _al("left")
        ws[f"B{row}"].border = bdr_L_med
        ws[f"C{row}"].value = 0
        ws[f"C{row}"].font  = _f(bold=True)
        ws[f"C{row}"].border = bdr_L_med_R_thin
        ws[f"C{row}"].number_format = '#,##0;-#,##0;"-   "'
        for dc in ("D","E","F","G","H","I"):
            ws[f"{dc}{row}"].value = 0
            ws[f"{dc}{row}"].font  = _f(color=GRAY)
            ws[f"{dc}{row}"].number_format = '#,##0;-#,##0;"-   "'
        ws[f"J{row}"].value = 0
        ws[f"J{row}"].font  = _f(color=GRAY)
        ws[f"J{row}"].border = bdr_R_med
        ws[f"J{row}"].number_format = '#,##0;-#,##0;"-   "'
        ws[f"L{row}"].value = 0
        ws[f"L{row}"].font  = _f(bold=True)
        ws[f"L{row}"].fill  = fill_usd
        if top_border:
            ws[f"L{row}"].border = _bdr(left=medium, right=thin, top=thin)
        else:
            ws[f"L{row}"].border = bdr_L_med_R_thin
        ws[f"L{row}"].number_format = '#,##0;-#,##0;"-   "'
        for uc in ("M","N","O","P","Q","R"):
            ws[f"{uc}{row}"].value = 0
            ws[f"{uc}{row}"].fill  = fill_usd
            ws[f"{uc}{row}"].alignment = _al("center")
            ws[f"{uc}{row}"].number_format = '#,##0;-#,##0;" -   "'
            if top_border:
                ws[f"{uc}{row}"].border = _bdr(top=thin)
        ws[f"S{row}"].value = 0
        ws[f"S{row}"].fill  = fill_usd
        ws[f"S{row}"].alignment = _al("center")
        ws[f"S{row}"].number_format = '#,##0;-#,##0;" -   "'
        if top_border:
            ws[f"S{row}"].border = _bdr(right=medium, top=thin)
        else:
            ws[f"S{row}"].border = bdr_R_med

    def _fila_blank(row, fill_usd=None):
        if fill_usd is None:
            fill_usd = fill_green
        ws[f"B{row}"].fill  = fill_white
        ws[f"B{row}"].border = bdr_L_med
        ws[f"C{row}"].border = bdr_L_med_R_thin
        ws[f"J{row}"].border = bdr_R_med
        ws[f"L{row}"].fill  = fill_usd
        ws[f"L{row}"].border = bdr_L_med_R_thin
        for uc in ("M","N","O","P","Q","R"):
            ws[f"{uc}{row}"].fill  = fill_usd
        ws[f"S{row}"].fill  = fill_usd
        ws[f"S{row}"].border = bdr_R_med

    def _fila_subtotal(row, label):
        ws[f"B{row}"].value = label
        ws[f"B{row}"].font  = _f(bold=True)
        ws[f"B{row}"].fill  = fill_white
        ws[f"B{row}"].border = bdr_L_med
        for col in ("C","D","E","F","G","H","I"):
            ws[f"{col}{row}"].value = 0
            ws[f"{col}{row}"].font  = _f(bold=True)
            ws[f"{col}{row}"].fill  = fill_orange
            ws[f"{col}{row}"].alignment = _al("center")
            ws[f"{col}{row}"].number_format = '#,##0;-#,##0;"-   "'
        ws[f"C{row}"].border = bdr_L_med_R_thin
        ws[f"J{row}"].value = 0
        ws[f"J{row}"].font  = _f(bold=True)
        ws[f"J{row}"].fill  = fill_orange
        ws[f"J{row}"].alignment = _al("center")
        ws[f"J{row}"].border = bdr_R_med
        ws[f"J{row}"].number_format = '#,##0;-#,##0;"-   "'
        for col in ("L","M","N","O","P","Q","R"):
            ws[f"{col}{row}"].value = 0
            ws[f"{col}{row}"].font  = _f(bold=True)
            ws[f"{col}{row}"].fill  = fill_orange
            ws[f"{col}{row}"].alignment = _al("center")
            ws[f"{col}{row}"].border = _bdr(top=thin, bottom=thin)
            ws[f"{col}{row}"].number_format = '#,##0;-#,##0;"-   "'
        ws[f"L{row}"].border = _bdr(left=medium, right=thin, top=thin, bottom=thin)
        ws[f"S{row}"].value = 0
        ws[f"S{row}"].font  = _f(bold=True)
        ws[f"S{row}"].fill  = fill_orange
        ws[f"S{row}"].alignment = _al("center")
        ws[f"S{row}"].border = _bdr(right=medium, top=thin, bottom=thin)
        ws[f"S{row}"].number_format = '#,##0;-#,##0;"-   "'

    # ── Filas 10-20: Materiales ──────────────────────────────────────────
    categorias = [
        (10, "DESINFECCION Y FERTILIZACION"),
        (11, "AMPLIACION "),
        (12, "CULTIVO TIERRA, CHAROLAS"),
        (13, "MATERIAL VEGETAL"),
        (14, "PREPARACION DE SUELO"),
        (15, "FERTILIZANTES (Manejo Integrado de Riego y Fertilizacion) "),
        (16, "DESINFECCION / PLAGUICIDAS (Manejo Integrado de Plagas y Enfermedades)"),
        (17, "MANTENIMIENTO"),
        (18, "EXPANSION CECILIA 25"),
        (19, "RENOVACION DE SIEMBRA"),
        (20, "MATERIAL DE EMPAQUE"),
    ]
    for i, (row, label) in enumerate(categorias):
        _fila_cat(row, label, top_border=(i == 0))
    _fila_blank(21)
    _fila_subtotal(22, "COSTO DE MATERIALES")
    _fila_blank(23)

    # ── Filas 24-59: Nominas ─────────────────────────────────────────────
    nominas = [
        (24, "NOMINA ADMON Oficina, Jefes de Finca, Ingenieros"),
        (25, "HORAS EXTR. DOM. Y FESTIVOS"),
        (26, "BONOS ASISIT, PUNTAULIDAD Y DESPENSA"),
        (27, "NOMINA PRODUCCION "),
        (28, "HORAS EXTR. DOM. Y FEST."),
        (29, "BONOS ASISIT, PUNT. Y DESP."),
        (30, "NOMINA PRODUCCION CORTE"),
        (31, "HORAS EXTR. DOM. Y FESTIVOS CORTE"),
        (32, "BONOS ASISIT, PUNTAULIDAD Y DESP. CORTE"),
        (33, "NOMINA PRODUCCION TRANSPLANTE"),
        (34, "HORAS EXTR. DOM. Y FEST. TRANSPLANTE"),
        (35, "BONOS ASISIT, PUNT. Y DESP. TRANSPLANTE"),
        (36, "NOMINA PRODUCCION MANEJO PLANTA"),
        (37, "HORAS EXTR. DOM. Y FEST. MANEJO PLANTA"),
        (38, "BONOS ASISIT, PUNT. Y DESP. MANEJO PLANTA"),
        (39, "NOMINA  HOOPS"),
        (40, "HORAS EXTR. DOM. Y FEST. HOOPS"),
        (41, "BONOS ASISIT, PUNT. Y DESP.HOOPS"),
        (42, "NOMINA  (MIPE,MIRFE,)"),
        (43, "HORAS EXTR. DOM. Y FEST. (MIPE,MIRFE)"),
        (44, "BONOS ASISIT, PUNT. Y DESP.(MIPE,MIRFE)"),
        (45, "NOMINA OPERATIVOS (TRACTORES, CAMEROS)"),
        (46, "HORAS EXTR. DOM. Y FEST. (TRACTORES, CAMEROS)"),
        (47, "BONOS ASISIT, PUNT. Y DESP. (TRACTORES, CAMEROS)"),
        (48, "NOMINA OPERATIVOS (CHOFER)"),
        (49, "HORAS EXTR. DOM. Y FEST. (CHOFER)"),
        (50, "BONOS ASISIT, PUNT. Y DESP. (CHOFER)"),
        (51, "NOMINA OPERATIVOS (VELADORES)"),
        (52, "HORAS EXTR. DOM. Y FEST. (VELADORES)"),
        (53, "BONOS ASISIT, PUNT. Y DESP. (VELADORES)"),
        (54, "NOMINA OPERATIVOS (SOLDADOR)"),
        (55, "HORAS EXTR. DOM. Y FEST. (SOLDADOR)"),
        (56, "BONOS ASISIT, PUNT. Y DESP. (SOLDADOR)"),
        (57, "NOMINA PRODUCCION Contratista y comisiones"),
        (58, "IMSS , INFONAVIT RCV"),
        (59, "1.8% al estado (1.2% tasa efectiva)"),
    ]
    for row, label in nominas:
        _fila_cat(row, label)
    _fila_blank(60)
    _fila_subtotal(61, "COSTO DE MANO DE OBRA")
    _fila_blank(62)

    # ── Filas 63-70: Servicios ───────────────────────────────────────────
    servicios = [
        (63, "ELECTRICIDAD"),
        (64, "FLETES Y ACARREOS (Flete aduana)"),
        (65, "GASTOS DE EXPORTACION "),
        (66, "CERTIFICADO DE FITOSANITARIOS"),
        (67, "Transporte de personal"),
        (68, "COMPRA DE FLOR A TERCEROS"),
        (69, "COMIDA PARA EL PERSONAL"),
        (70, "RO, TEL, RTA.ALIM."),
    ]
    for row, label in servicios:
        _fila_cat(row, label)
    _fila_blank(71)
    _fila_subtotal(72, "COSTO DE SERVICIOS")

    # Fila 73: separadora con bordes top/bottom
    ws["B73"].fill  = fill_white
    ws["B73"].border = _bdr(left=medium, top=thin, bottom=thin)
    ws["C73"].border = _bdr(left=medium, top=thin, bottom=thin)
    ws["L73"].fill  = fill_green; ws["L73"].border = bdr_L_med
    for col in ("M","N","O","P","Q","R"):
        ws[f"{col}73"].fill = fill_green
    ws["S73"].fill  = fill_green; ws["S73"].border = bdr_R_med

    # ── Fila 74: COSTO DE PRODUCCION Y VENTAS ────────────────────────────
    ws["B74"].value = "COSTO DE PRODUCCION Y VENTAS"
    ws["B74"].font  = _f(bold=True)
    ws["B74"].fill  = fill_white
    ws["B74"].border = _bdr(left=medium, bottom=medium)
    for col in ("D","E","F","G","H","I"):
        ws[f"{col}74"].value = 0
        ws[f"{col}74"].font  = _f(bold=True)
        ws[f"{col}74"].border = _bdr(bottom=medium)
        ws[f"{col}74"].number_format = '#,##0;-#,##0;"-   "'
    ws["C74"].value = 0; ws["C74"].font = _f(bold=True)
    ws["C74"].border = _bdr(left=medium, right=thin, top=thin, bottom=medium)
    ws["C74"].number_format = '#,##0;-#,##0;"-   "'
    ws["J74"].value = 0; ws["J74"].font = _f(bold=True)
    ws["J74"].border = _bdr(right=medium, bottom=medium)
    ws["J74"].number_format = '#,##0;-#,##0;"-   "'
    for col in ("M","N","O","P","Q","R"):
        ws[f"{col}74"].value = 0; ws[f"{col}74"].font = _f(bold=True)
        ws[f"{col}74"].fill  = fill_green
        ws[f"{col}74"].border = _bdr(top=thin, bottom=medium)
        ws[f"{col}74"].number_format = '#,##0;-#,##0;"-   "'
    ws["L74"].value = 0; ws["L74"].font = _f(bold=True)
    ws["L74"].fill  = fill_green
    ws["L74"].border = _bdr(left=medium, right=thin, top=thin, bottom=medium)
    ws["L74"].number_format = '#,##0;-#,##0;"-   "'
    ws["S74"].value = 0; ws["S74"].font = _f(bold=True)
    ws["S74"].fill  = fill_green
    ws["S74"].border = _bdr(right=medium, top=thin, bottom=medium)
    ws["S74"].number_format = '#,##0;-#,##0;"-   "'
    ws.row_dimensions[74].height = 15

    # ── Filas 76-92: Produccion ───────────────────────────────────────────
    produccion = [
        (76, "CAJAS PROCESADAS TOTALES"),
        (77, "INVENTARIO INICIAL"),
        (78, "TALLOS COSECHADOS"),
        (79, "TALLOS DESECHADOS"),
        (80, "TALLOS DESECHADOS sf"),
        (81, "TALLOS COMPRADOS"),
        (82, "TALLOS EN BOUQUETS O PROCESADOS"),
        (83, "TALLOS DESPACHADOS"),
        (84, "LIBRAS DESPACHADAS ALBAHACA"),
        (85, "TALLOS muestra"),
        (86, "INVENTARIO FINAL"),
        (87, "TALLOS PROCESADOS TOTALES"),
        (88, " CHAROLAS SEMBRADAS *288 PLUGS ="),
        (89, " NUMERO DE CHAROLAS SEMBRADAS "),
        (90, " NUMERO DE ESQUEJES SEMBRADOS"),
        (91, " METROS DE SIEMBRA"),
        (92, " HECTAREAS EN SIEMBRA"),
    ]
    for i, (row, label) in enumerate(produccion):
        first = (i == 0)
        last  = (i == len(produccion)-1)
        ws[f"B{row}"].value = label
        ws[f"B{row}"].fill  = fill_white
        ws[f"B{row}"].alignment = _al("left")
        b_bdr = _bdr(left=medium, top=medium) if first else (_bdr(left=medium, bottom=medium) if last else bdr_L_med)
        ws[f"B{row}"].border = b_bdr
        ws[f"C{row}"].value = 0
        ws[f"C{row}"].font  = _f(bold=True)
        c_bdr = _bdr(left=medium, right=thin, top=medium) if first else (_bdr(left=medium, right=thin, bottom=medium) if last else bdr_L_med_R_thin)
        ws[f"C{row}"].border = c_bdr
        for dc in ("D","E","F","G","H","I"):
            ws[f"{dc}{row}"].value = 0
            ws[f"{dc}{row}"].border = _bdr(top=medium) if first else (_bdr(bottom=medium) if last else _bdr())
        ws[f"J{row}"].value = 0
        ws[f"J{row}"].border = _bdr(right=medium, top=medium) if first else (_bdr(right=medium, bottom=medium) if last else bdr_R_med)
        ws[f"L{row}"].value = 0
        ws[f"L{row}"].font  = _f(bold=True)
        ws[f"L{row}"].fill  = fill_yellow
        ws[f"L{row}"].border = _bdr(left=medium, right=thin, top=medium) if first else (_bdr(left=medium, right=thin, bottom=medium) if last else bdr_L_med_R_thin)
        for uc in ("M","N","O","P","Q","R"):
            ws[f"{uc}{row}"].value = 0
            ws[f"{uc}{row}"].fill  = fill_yellow
            ws[f"{uc}{row}"].alignment = _al("center")
            ws[f"{uc}{row}"].border = _bdr(top=medium) if first else (_bdr(bottom=medium) if last else _bdr())
        ws[f"S{row}"].value = 0
        ws[f"S{row}"].fill  = fill_yellow
        ws[f"S{row}"].alignment = _al("center")
        ws[f"S{row}"].border = _bdr(right=medium, top=medium) if first else (_bdr(right=medium, bottom=medium) if last else bdr_R_med)

    ws.row_dimensions[92].height = 15

    # ── Fila 93 ───────────────────────────────────────────────────────────
    ws["B93"].value = "<<< INDICADORES"
    ws["B93"].font  = _f(bold=True)
    ws.row_dimensions[93].height = 15

    # ── Filas 94-121: Costos unitarios ────────────────────────────────────
    ws["B94"].value = "COSTOS UNITARIOS"; ws["B94"].font = _f(bold=True)
    ws["B94"].border = _bdr(left=medium, top=medium)
    ws["L94"].fill  = fill_green; ws["L94"].border = _bdr(left=medium, right=thin, top=medium)
    for col in ("M","N","O","P","Q","R"): ws[f"{col}94"].fill = fill_green; ws[f"{col}94"].border = _bdr(top=medium)
    ws["S94"].fill  = fill_green; ws["S94"].border = _bdr(right=medium, top=medium)

    def _cu_row(row, label, bold=False, fill_b=None):
        ws[f"B{row}"].value = label; ws[f"B{row}"].font = _f(bold=bold)
        if fill_b: ws[f"B{row}"].fill = fill_b
        ws[f"B{row}"].border = bdr_L_med
        ws[f"C{row}"].value = 0; ws[f"C{row}"].font = _f(bold=True)
        ws[f"C{row}"].border = bdr_L_med_R_thin
        ws[f"L{row}"].value = 0; ws[f"L{row}"].font = _f(bold=True)
        ws[f"L{row}"].fill  = fill_green; ws[f"L{row}"].border = bdr_L_med_R_thin
        for col in ("M","N","O","P","Q","R"): ws[f"{col}{row}"].fill = fill_green
        ws[f"S{row}"].fill = fill_green; ws[f"S{row}"].border = bdr_R_med
        if fill_b:
            ws[f"C{row}"].fill = fill_b
            for col in ("L","M","N","O","P","Q","R","S"): ws[f"{col}{row}"].fill = fill_b

    _cu_row(95, "$ / Tallo Procesado", bold=True)
    _cu_row(96, "COSTOS UNITARIOS", bold=True)
    _cu_row(97, "$ / Libras Procesadas", bold=True)
    ws["L97"].border = _bdr(left=medium, right=thin, bottom=thin)

    _cu_row(98, "Materiales")
    ws["B98"].border = _bdr(left=medium, top=thin)
    ws["C98"].border = _bdr(left=medium, right=thin, top=thin)
    ws["L98"].border = _bdr(left=medium, right=thin, top=thin)
    for col in ("M","N","O","P","Q","R"): ws[f"{col}98"].border = _bdr(top=thin)
    ws["S98"].border = _bdr(right=medium, top=thin)

    _cu_row(99, "Mano de Obra")
    _cu_row(100, "Servicios (Fletes)")
    ws["B100"].border = _bdr(left=medium, bottom=thin)
    ws["C100"].border = _bdr(left=medium, right=thin, bottom=thin)
    ws["L100"].border = _bdr(left=medium, right=thin, bottom=thin)
    for col in ("M","N","O","P","Q","R"): ws[f"{col}100"].border = _bdr(bottom=thin)
    ws["S100"].border = _bdr(right=medium, bottom=thin)

    _cu_row(101, "Costo de Produccion y Ventas", bold=True, fill_b=fill_orange)
    ws["B101"].border = _bdr(left=medium, top=thin, bottom=thin)
    ws["C101"].border = _bdr(left=medium, right=thin, top=thin, bottom=thin)
    ws["L101"].border = _bdr(left=medium, right=thin, bottom=thin)
    ws["S101"].border = _bdr(right=medium, bottom=thin)

    # Spacers 102, 104, 107
    for row in (102, 104, 107):
        ws[f"C{row}"].border = bdr_L_med_R_thin
        ws[f"L{row}"].fill = fill_green; ws[f"L{row}"].border = bdr_L_med_R_thin
        for col in ("M","N","O","P","Q","R"): ws[f"{col}{row}"].fill = fill_green
        ws[f"S{row}"].fill = fill_green; ws[f"S{row}"].border = bdr_R_med

    _cu_row(103, "Material de Empaque / Tallo", bold=True)
    ws["B103"].border = _bdr(left=medium, top=thin, bottom=thin)
    ws["C103"].border = _bdr(left=medium, right=thin, top=thin, bottom=thin)
    ws["L103"].border = _bdr(left=medium, right=thin, top=thin, bottom=thin)
    for col in ("M","N","O","P","Q","R"): ws[f"{col}103"].border = _bdr(top=thin, bottom=thin)
    ws["S103"].border = _bdr(right=medium, top=thin, bottom=thin)

    _cu_row(105, "Sanidad Vegetal / Tallo", bold=True)
    ws["B105"].border = _bdr(left=medium, top=thin)
    ws["C105"].border = _bdr(left=medium, right=thin, top=thin)
    ws["L105"].border = _bdr(left=medium, right=thin, top=thin)
    for col in ("M","N","O","P","Q","R"): ws[f"{col}105"].border = _bdr(top=thin)
    ws["S105"].border = _bdr(right=medium, top=thin)

    _cu_row(106, "Fertlizacion / Tallo", bold=True)
    ws["B106"].border = _bdr(left=medium, bottom=thin)
    ws["C106"].border = _bdr(left=medium, right=thin, bottom=thin)
    ws["L106"].border = _bdr(left=medium, right=thin, bottom=thin)
    for col in ("M","N","O","P","Q","R"): ws[f"{col}106"].border = _bdr(bottom=thin)
    ws["S106"].border = _bdr(right=medium, bottom=thin)

    _cu_row(108, "Mano de Obra Prod / Tallo", bold=True)
    ws["B108"].border = _bdr(left=medium, top=thin, bottom=medium)
    ws["C108"].border = _bdr(left=medium, right=thin, top=thin, bottom=medium)
    ws["L108"].border = _bdr(left=medium, right=thin, top=thin, bottom=medium)
    for col in ("M","N","O","P","Q","R"): ws[f"{col}108"].border = _bdr(top=thin, bottom=medium)
    ws["S108"].border = _bdr(right=medium, top=thin, bottom=medium)
    ws.row_dimensions[108].height = 15
    ws.row_dimensions[109].height = 15

    # ── Fila 110-121: $ / Hectarea ────────────────────────────────────────
    ws["B110"].value = "$ / Hectarea"; ws["B110"].font = _f(bold=True)
    ws["B110"].border = _bdr(left=medium, top=medium)
    ws["C110"].border = _bdr(left=medium, right=thin, top=medium, bottom=thin)
    ws["J110"].border = _bdr(right=medium, top=medium)
    ws["L110"].fill = fill_yellow; ws["L110"].border = _bdr(left=medium, right=thin, top=medium)
    for col in ("M","N","O","P","Q","R"):
        ws[f"{col}110"].fill = fill_yellow; ws[f"{col}110"].border = _bdr(top=medium)
    ws["S110"].fill = fill_yellow; ws["S110"].border = _bdr(right=medium, top=medium)

    def _ha_row(row, label, top_b=False, bottom_b=False, both_b=False):
        ws[f"B{row}"].value = label; ws[f"B{row}"].font = _f()
        ws[f"B{row}"].fill  = fill_white; ws[f"B{row}"].alignment = _al("left")
        if both_b:   ws[f"B{row}"].border = _bdr(left=medium, top=thin, bottom=thin)
        elif top_b:  ws[f"B{row}"].border = _bdr(left=medium, top=thin)
        elif bottom_b: ws[f"B{row}"].border = _bdr(left=medium, bottom=thin)
        else:        ws[f"B{row}"].border = bdr_L_med
        ws[f"C{row}"].value = 0; ws[f"C{row}"].font = _f(bold=True)
        ws[f"C{row}"].border = _bdr(left=medium, right=thin,
                                    top=(thin if top_b or both_b else none_s),
                                    bottom=(thin if bottom_b or both_b else none_s))
        ws[f"J{row}"].border = bdr_R_med
        ws[f"L{row}"].value = 0; ws[f"L{row}"].font = _f(bold=True)
        ws[f"L{row}"].fill  = fill_yellow
        ws[f"L{row}"].border = _bdr(left=medium, right=thin,
                                    top=(thin if top_b or both_b else none_s),
                                    bottom=(thin if bottom_b or both_b else none_s))
        for col in ("M","N","O","P","Q","R"):
            ws[f"{col}{row}"].fill = fill_yellow
            ws[f"{col}{row}"].border = _bdr(top=(thin if top_b or both_b else none_s),
                                            bottom=(thin if bottom_b or both_b else none_s))
        ws[f"S{row}"].fill = fill_yellow
        ws[f"S{row}"].border = _bdr(right=medium,
                                    top=(thin if top_b or both_b else none_s),
                                    bottom=(thin if bottom_b or both_b else none_s))

    _ha_row(111, "Materiales", top_b=True)
    _ha_row(112, "Mano de Obra")
    _ha_row(113, "Servicios (Fletes)", bottom_b=True)
    # spacer 115
    ws[f"C115"].border = bdr_L_med_R_thin
    ws["L115"].fill = fill_yellow; ws["L115"].border = bdr_L_med_R_thin
    for col in ("M","N","O","P","Q","R"): ws[f"{col}115"].fill = fill_yellow
    ws["S115"].fill = fill_yellow; ws["S115"].border = bdr_R_med

    _ha_row(114, "Costo de Produccion y Ventas", both_b=True)
    _ha_row(116, "Material de Empaque / Caja", both_b=True)
    ws["B116"].font = _f(bold=True)
    # spacer 117
    ws["C117"].border = bdr_L_med_R_thin
    ws["L117"].fill = fill_yellow; ws["L117"].border = bdr_L_med_R_thin
    for col in ("M","N","O","P","Q","R"): ws[f"{col}117"].fill = fill_yellow
    ws["S117"].fill = fill_yellow; ws["S117"].border = bdr_R_med

    _ha_row(118, "Sanidad Vegetal / Ha", top_b=True)
    ws["B118"].font = _f(bold=True)
    _ha_row(119, "Fertlizacion / Ha", bottom_b=True)
    ws["B119"].font = _f(bold=True)
    # spacer 120
    ws["C120"].border = bdr_L_med_R_thin
    ws["L120"].fill = fill_yellow; ws["L120"].border = bdr_L_med_R_thin
    for col in ("M","N","O","P","Q","R"): ws[f"{col}120"].fill = fill_yellow
    ws["S120"].fill = fill_yellow; ws["S120"].border = bdr_R_med

    ws.row_dimensions[121].height = 15
    ws["B121"].value = "Mano de Obra Prod / Ha"; ws["B121"].font = _f(bold=True)
    ws["B121"].fill = fill_white
    ws["B121"].border = _bdr(left=medium, top=thin, bottom=medium)
    ws["C121"].value = 0; ws["C121"].font = _f(bold=True)
    ws["C121"].border = _bdr(left=medium, right=thin, top=thin, bottom=medium)
    ws["L121"].value = 0; ws["L121"].font = _f(bold=True)
    ws["L121"].fill = fill_yellow
    ws["L121"].border = _bdr(left=medium, right=thin, top=thin, bottom=medium)
    for col in ("M","N","O","P","Q","R"):
        ws[f"{col}121"].fill = fill_yellow
        ws[f"{col}121"].border = _bdr(top=thin, bottom=medium)
    ws["S121"].fill = fill_yellow
    ws["S121"].border = _bdr(right=medium, top=thin, bottom=medium)

    # ── KPI's ─────────────────────────────────────────────────────────────
    ws["B124"].value = "KPI's "; ws["B124"].font = _f(bold=True)

    # Proyectos de inversion
    ws["B125"].value = "Proyectos de inversion"
    ws["B125"].font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
    ws["B125"].fill  = fill_kpi
    ws["B125"].alignment = _al("left")
    ws["B125"].border = _bdr(left=thin, right=thin, top=thin)
    ws["L125"].value = "Total Weekly"
    ws["L125"].font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
    ws["L125"].fill  = fill_kpi
    ws["L125"].alignment = _al("center")
    ws["L125"].border = _bdr(left=thin, right=thin, top=thin, bottom=thin)

    proyectos = [
        (126, "Sistema de riego (Ramona)"),
        (127, "Sistema de riego (Isabella)"),
        (128, "Caseta (Isabella)"),
        (129, "Sistema de ventilacion"),
        (130, "Sistema de tratamiento de aguas residuales (Isabella)"),
        (131, "Arcos para invernaderos "),
        (132, "proyecto luz"),
        (133, "Construccion de Almacen (Ramona) "),
        (134, "Construccion de Almacen (Isabela) "),
        (135, "Carritos"),
        (136, "Maquinaria "),
        (137, "Chiller"),
        (138, "Cuarto frio"),
        (139, "veronicas"),
    ]
    for row, label in proyectos:
        ws[f"B{row}"].value = label
        ws[f"B{row}"].fill  = fill_white
        ws[f"B{row}"].border = _bdr(left=thin, right=thin)
        ws[f"C{row}"].value = 0
        ws[f"C{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
        ws[f"C{row}"].fill  = fill_white
        ws[f"C{row}"].border = _bdr(left=thin, right=thin)
        ws[f"C{row}"].number_format = '"$"#,##0;-"$"#,##0;" $-   "'
        ws[f"J{row}"].border = _bdr(right=thin)
        ws[f"L{row}"].value = 0
        ws[f"L{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
        ws[f"L{row}"].fill  = fill_white
        ws[f"L{row}"].border = _bdr(left=thin, right=thin)
        ws[f"L{row}"].number_format = '" $"#,##0;-" $"#,##0;" $-   "'
        for uc in ("M","N","O","P","Q","R","S"):
            ws[f"{uc}{row}"].value = 0
            ws[f"{uc}{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
            ws[f"{uc}{row}"].fill  = fill_white
            ws[f"{uc}{row}"].border = _bdr(left=thin, right=thin)

    ws["B139"].border = _bdr(left=thin, right=thin, bottom=thin)
    ws["C139"].border = _bdr(left=thin, right=thin, bottom=thin)
    ws["J139"].border = _bdr(right=thin, bottom=thin)
    for uc in ("L","M","N","O","P","Q","R","S"):
        ws[f"{uc}139"].border = _bdr(left=thin, right=thin, bottom=thin)

    ws["B140"].value = "Total "
    ws["B140"].font  = _f(bold=True)
    ws["B140"].fill  = fill_white
    ws["B140"].border = _bdr(left=thin, right=thin, top=thin, bottom=thin)
    ws["C140"].value = 0
    ws["C140"].font  = Font(color="0000FF", name="Calibri", size=10)
    ws["C140"].border = _bdr(top=thin, bottom=thin)
    ws["C140"].number_format = '" $"#,##0;-" $"#,##0;" $-   "'
    ws["L140"].value = 0
    ws["L140"].font  = Font(color="0000FF", name="Calibri", size=10)
    ws["L140"].border = _bdr(left=thin, right=thin, bottom=thin)
    ws["L140"].number_format = '" $"#,##0;-" $"#,##0;" $-   "'

    # Logistica
    ws["B143"].value = "Logistica "
    ws["B143"].font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
    ws["B143"].fill  = fill_kpi
    ws["B143"].alignment = _al("left")
    ws["B143"].border = _bdr(left=thin, top=thin)
    ws["J143"].border = _bdr(right=thin, top=thin)
    ws["L143"].value = "Total Weekly"
    ws["L143"].font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
    ws["L143"].fill  = fill_kpi
    ws["L143"].alignment = _al("center")
    ws["L143"].border = _bdr(left=thin, right=thin, top=thin, bottom=thin)
    ws["N143"].value = "PosCo-RM"
    ws["N143"].font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
    ws["N143"].fill  = fill_kpi
    ws["N143"].alignment = _al("center")
    ws["N143"].border = _bdr(left=thin, right=thin, top=thin, bottom=thin)

    logistica = [
        (144, "Numero de camiones despachados "),
        (145, "Numero de tarimas despachadas (montadas al camion)"),
        (146, "Numero de cajas despachadas"),
        (147, "Numero de Pies cubicos de cajas despachadas "),
        (148, "Numero de Pies cubicos promedio / camion despachado "),
        (149, "Capacidad en pies cubicos por camion "),
        (150, "Rendimiento promedio por camion "),
    ]
    for row, label in logistica:
        ws[f"B{row}"].value = label
        ws[f"B{row}"].fill  = fill_white
        ws[f"B{row}"].border = _bdr(left=thin)
        ws[f"C{row}"].value = 0
        ws[f"C{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
        ws[f"C{row}"].fill  = fill_white
        ws[f"C{row}"].border = _bdr(left=thin, right=thin)
        ws[f"J{row}"].border = _bdr(right=thin)
        ws[f"L{row}"].value = 0
        ws[f"L{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
        ws[f"L{row}"].fill  = fill_white
        ws[f"L{row}"].border = _bdr(left=thin, right=thin)
        ws[f"N{row}"].value = 0
        ws[f"N{row}"].font  = _f(bold=True)
        ws[f"N{row}"].fill  = fill_white
        ws[f"N{row}"].border = _bdr(right=thin)
        ws[f"S{row}"].border = _bdr(right=thin)

    kpi_groups = [
        (152, "Costo incurrido por flete, gtos expo, fitosanitarios"),
        (153, "Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
        (154, "Numero de Camiones despachados "),
        (156, "Costo incurrido promedio flete, gtos expo, fitosanitarios / pie cubico"),
        (157, "Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
        (158, "Numero de Pies cubicos de cajas despachadas"),
        (160, "Costo incurrido flete, gtos expo, fitosanitarios / cajas despachadas"),
        (161, "Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
        (162, "Numero de cajas despachadas"),
    ]
    for row, label in kpi_groups:
        ws[f"B{row}"].value = label
        ws[f"B{row}"].fill  = fill_white
        ws[f"B{row}"].border = _bdr(left=thin)
        ws[f"C{row}"].value = 0
        ws[f"C{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
        ws[f"C{row}"].fill  = fill_white
        ws[f"C{row}"].border = _bdr(left=thin, right=thin)
        ws[f"L{row}"].value = 0
        ws[f"L{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
        ws[f"L{row}"].fill  = fill_white
        ws[f"L{row}"].border = _bdr(left=thin, right=thin)
        ws[f"N{row}"].value = 0
        ws[f"N{row}"].border = _bdr(right=thin)

    ws["B165"].value = "Material de empaque / Caja"
    ws["B165"].font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
    ws["B165"].fill  = fill_kpi
    ws["B165"].alignment = _al("left")
    ws["B165"].border = _bdr(left=thin, top=thin)

    me_rows = [
        (166, "Costo incurrido en Material de empaque / pie cubico"),
        (167, "Costo incurrido en Material de empaque (USD)"),
        (168, "Numero de Pies cubicos de cajas despachadas"),
        (170, "Costo incurrido en Material de empaque / cajas despachadas"),
        (171, "Costo incurrido en Material de empaque (USD)"),
        (172, "Numero de cajas despachadas"),
    ]
    for row, label in me_rows:
        ws[f"B{row}"].value = label
        ws[f"B{row}"].fill  = fill_white
        ws[f"B{row}"].border = _bdr(left=thin)
        ws[f"C{row}"].value = 0
        ws[f"C{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
        ws[f"C{row}"].fill  = fill_white
        ws[f"C{row}"].border = _bdr(left=thin, right=thin)
        ws[f"L{row}"].value = 0
        ws[f"L{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
        ws[f"L{row}"].fill  = fill_white
        ws[f"L{row}"].border = _bdr(left=thin, right=thin)
        ws[f"N{row}"].value = 0
        ws[f"N{row}"].border = _bdr(right=thin)

    # Merged cells
    merges = [
        "C5:J5", "L5:R5",
        "C153:C154", "L153:L154",
        "C157:C158", "L157:L158",
        "C161:C162", "L161:L162",
        "C167:C168", "L167:L168",
        "C171:C172", "L171:L172",
    ]
    for m in merges:
        try:
            ws.merge_cells(m)
        except Exception:
            pass

# --- Crear nueva hoja WK en SharePoint via Microsoft Graph API (con sesión) ---
def crear_hoja_wk(nombre_hoja: str, tenant_id: str, client_id: str, client_secret: str) -> dict:
    """
    Crea una nueva hoja WK#### desde cero usando una sesión de workbook de Graph API.
    Funciona aunque el archivo esté abierto por otros usuarios (no requiere lock).
    Escribe todas las celdas directamente via API, sin subir el archivo completo.
    Requiere Files.ReadWrite en la App Registration de Azure AD.
    """
    import base64 as _b64
    import time

    # ── Helper: construir lista plana de celdas { address, value } ────────
    def _celdas_de_la_hoja(nombre):
        """Devuelve lista de dicts con address y valor para escribir via Graph API."""
        celdas = []

        def c(addr, val):
            celdas.append({"address": addr, "value": val})

        # Encabezados
        c("B1", "CENTRO FLORICULTOR DE BAJA CALIFORNIA, S.A. DE C.V. ")
        c("B2", "SEMANA DE CALCULO - Mexico")
        c("B3", nombre)
        c("C3", 19); c("D3", " tipo de cambio")
        c("L3", 19); c("M3", "  tipo de cambio ")
        c("B4", "Del ___ al ___ de ________ 20__")
        c("C5", "(MXN) Pesos Mexicanos")
        c("L5", "US Dollars")
        c("B6", "TOTAL FINCA")
        # Fila 7: ranchos MXN
        for col, h in zip(["C","D","E","F","G","H","I","J"],
                          ["TOTAL","Prop-RM","PosCo-RM","Campo -RM","Isabela","Christina","Cecilia","Cecilia 25"]):
            c(f"{col}7", h)
        # Fila 7: ranchos USD
        for col, h in zip(["L","M","N","O","P","Q","R","S"],
                          ["TOTAL","Prop-RM","PosCo-RM","Campo -RM","ISABELA","Christina","CECILIA","CECILIA 25"]):
            c(f"{col}7", h)
        c("B7", "Produccion"); c("C8", "SEMANAL "); c("L8", '"WEEKLY"')
        c("B9", "EJECUCION SEMANAL")
        for col in ["D","E","F","G","H","I","J"]:
            c(f"{col}9", 1)

        # Categorías de materiales (filas 10-20)
        categorias = [
            (10, "DESINFECCION Y FERTILIZACION"),
            (11, "AMPLIACION "),
            (12, "CULTIVO TIERRA, CHAROLAS"),
            (13, "MATERIAL VEGETAL"),
            (14, "PREPARACION DE SUELO"),
            (15, "FERTILIZANTES (Manejo Integrado de Riego y Fertilización) "),
            (16, "DESINFECCION / PLAGUICIDAS (Manejo Integrado de Plagas y Enfermedades)"),
            (17, "MANTENIMIENTO"),
            (18, "EXPANSION CECILIA 25"),
            (19, "RENOVACION DE SIEMBRA"),
            (20, "MATERIAL DE EMPAQUE"),
        ]
        for row, label in categorias:
            c(f"B{row}", label)

        c("B22", "COSTO DE MATERIALES")

        # Nóminas (filas 24-59)
        nominas = [
            (24, "NOMINA ADMON Oficina, Jefes de Finca, Ingenieros"),
            (25, "HORAS EXTR. DOM. Y FESTIVOS"),
            (26, "BONOS ASISIT, PUNTAULIDAD Y DESPENSA"),
            (27, "NOMINA PRODUCCION "),
            (28, "HORAS EXTR. DOM. Y FEST."),
            (29, "BONOS ASISIT, PUNT. Y DESP."),
            (30, "NOMINA PRODUCCION CORTE"),
            (31, "HORAS EXTR. DOM. Y FESTIVOS CORTE"),
            (32, "BONOS ASISIT, PUNTAULIDAD Y DESP. CORTE"),
            (33, "NOMINA PRODUCCION TRANSPLANTE"),
            (34, "HORAS EXTR. DOM. Y FEST. TRANSPLANTE"),
            (35, "BONOS ASISIT, PUNT. Y DESP. TRANSPLANTE"),
            (36, "NOMINA PRODUCCION MANEJO PLANTA"),
            (37, "HORAS EXTR. DOM. Y FEST. MANEJO PLANTA"),
            (38, "BONOS ASISIT, PUNT. Y DESP. MANEJO PLANTA"),
            (39, "NOMINA  HOOPS"),
            (40, "HORAS EXTR. DOM. Y FEST. HOOPS"),
            (41, "BONOS ASISIT, PUNT. Y DESP.HOOPS"),
            (42, "NOMINA  (MIPE,MIRFE,)"),
            (43, "HORAS EXTR. DOM. Y FEST. (MIPE,MIRFE)"),
            (44, "BONOS ASISIT, PUNT. Y DESP.(MIPE,MIRFE)"),
            (45, "NOMINA OPERATIVOS (TRACTORES, CAMEROS)"),
            (46, "HORAS EXTR. DOM. Y FEST. (TRACTORES, CAMEROS)"),
            (47, "BONOS ASISIT, PUNT. Y DESP. (TRACTORES, CAMEROS)"),
            (48, "NOMINA OPERATIVOS (CHOFER)"),
            (49, "HORAS EXTR. DOM. Y FEST. (CHOFER)"),
            (50, "BONOS ASISIT, PUNT. Y DESP. (CHOFER)"),
            (51, "NOMINA OPERATIVOS (VELADORES)"),
            (52, "HORAS EXTR. DOM. Y FEST. (VELADORES)"),
            (53, "BONOS ASISIT, PUNT. Y DESP. (VELADORES)"),
            (54, "NOMINA OPERATIVOS (SOLDADOR)"),
            (55, "HORAS EXTR. DOM. Y FEST. (SOLDADOR)"),
            (56, "BONOS ASISIT, PUNT. Y DESP. (SOLDADOR)"),
            (57, "NOMINA PRODUCCION Contratista y comisiones"),
            (58, "IMSS , INFONAVIT RCV"),
            (59, "1.8% al estado (1.2% tasa efectiva)"),
        ]
        for row, label in nominas:
            c(f"B{row}", label)

        c("B61", "COSTO DE MANO DE OBRA")

        # Servicios (filas 63-70)
        servicios = [
            (63, "ELECTRICIDAD"),
            (64, "FLETES Y ACARREOS (Flete aduana)"),
            (65, "GASTOS DE EXPORTACION "),
            (66, "CERTIFICADO DE FITOSANITARIOS"),
            (67, "Transporte de personal"),
            (68, "COMPRA DE FLOR A TERCEROS"),
            (69, "COMIDA PARA EL PERSONAL"),
            (70, "RO, TEL, RTA.ALIM."),
        ]
        for row, label in servicios:
            c(f"B{row}", label)

        c("B72", "COSTO DE SERVICIOS")
        c("B74", "COSTO DE PRODUCCION Y VENTAS")

        # Producción (filas 76-92)
        prod = [
            (76, "CAJAS PROCESADAS TOTALES"),
            (77, "INVENTARIO INICIAL"),
            (78, "TALLOS COSECHADOS"),
            (79, "TALLOS DESECHADOS"),
            (80, "TALLOS DESECHADOS sf"),
            (81, "TALLOS COMPRADOS"),
            (82, "TALLOS EN BOUQUETS O PROCESADOS"),
            (83, "TALLOS DESPACHADOS"),
            (84, "LIBRAS DESPACHADAS ALBAHACA"),
            (85, "TALLOS muestra"),
            (86, "INVENTARIO FINAL"),
            (87, "TALLOS PROCESADOS TOTALES"),
            (88, " CHAROLAS SEMBRADAS *288 PLUGS ="),
            (89, " NUMERO DE CHAROLAS SEMBRADAS "),
            (90, " NUMERO DE ESQUEJES SEMBRADOS"),
            (91, " METROS DE SIEMBRA"),
            (92, " HECTAREAS EN SIEMBRA"),
        ]
        for row, label in prod:
            c(f"B{row}", label)

        c("B93", "<<< INDICADORES")
        c("B94", "COSTOS UNITARIOS"); c("B95", "$ / Tallo Procesado")
        c("B96", "COSTOS UNITARIOS"); c("B97", "$ / Libras Procesadas")
        for row, label in [(98,"Materiales"),(99,"Mano de Obra"),(100,"Servicios (Fletes)"),
                           (101,"Costo de Produccion y Ventas"),(103,"Material de Empaque / Tallo"),
                           (105,"Sanidad Vegetal / Tallo"),(106,"Fertlizacion / Tallo"),
                           (108,"Mano de Obra Prod / Tallo")]:
            c(f"B{row}", label)
        c("B110", "$ / Hectarea")
        for row, label in [(111,"Materiales"),(112,"Mano de Obra"),(113,"Servicios (Fletes)"),
                           (114,"Costo de Produccion y Ventas"),(121,"Mano de Obra Prod / Ha")]:
            c(f"B{row}", label)

        c("B124", "KPI's ")
        c("B125", "Proyectos de inversión"); c("L125", "Total Weekly")
        proyectos = [
            (126,"Sistema de riego (Ramona)"),(127,"Sistema de riego (Isabella)"),
            (128,"Caseta (Isabella)"),(129,"Sistema de ventilacion)"),
            (130,"Sistema de tratamiento de aguas residuales (Isabella)"),
            (131,"Arcos para invernaderos "),(132,"proyecto luz"),
            (133,"Construcción de Almacén (Ramona) "),(134,"Construcción de Almacén (Isabela) "),
            (135,"Carritos"),(136,"Maquinaria "),(137,"Chiller"),
            (138,"Cuarto frio"),(139,"veronicas"),
        ]
        for row, label in proyectos:
            c(f"B{row}", label)
        c("B140", "Total ")
        c("B143", "Logística "); c("L143", "Total Weekly"); c("N143", "PosCo-RM")
        for row, label in [
            (144,"Número de camiones despachados "),(145,"Número de tarimas despachadas (montadas al camión)"),
            (146,"Número de cajas despachadas"),(147,"Número de Pies cúbicos de cajas despachadas "),
            (148,"Número de Pies cubicos promedio / camión despachado "),
            (149,"Capacidad en pies cúbicos por camión "),(150,"Rendimiento promedio por camión "),
        ]:
            c(f"B{row}", label)
        for row, label in [
            (152,"Costo incurrido por flete, gtos expo, fitosanitarios"),
            (153,"Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
            (154,"Número de Camiones despachados "),
            (156,"Costo incurrido promedio flete, gtos expo, fitosanitarios / pie cúbico"),
            (157,"Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
            (158,"Número de Pies cúbicos de cajas despachadas"),
            (160,"Costo incurrido flete, gtos expo, fitosanitarios / cajas despachadas"),
            (161,"Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
            (162,"Número de cajas despachadas"),
        ]:
            c(f"B{row}", label)
        c("B165", "Material de empaque / Caja")
        for row, label in [
            (166,"Costo incurrido en Material de empaque / pie cúbico"),
            (167,"Costo incurrido en Material de empaque (USD)"),
            (168,"Número de Pies cúbicos de cajas despachadas"),
            (170,"Costo incurrido en Material de empaque / cajas despachadas"),
            (171,"Costo incurrido en Material de empaque (USD)"),
            (172,"Número de cajas despachadas"),
        ]:
            c(f"B{row}", label)

        return celdas

    # ── 1. Token OAuth2 ───────────────────────────────────────────────────
    token_url = f'https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token'
    token_resp = requests.post(token_url, data={
        "grant_type":    "client_credentials",
        "client_id":     client_id,
        "client_secret": client_secret,
        "scope":         "https://graph.microsoft.com/.default",
    }, timeout=20)
    if token_resp.status_code != 200:
        return {"ok": False, "error": f"Error obteniendo token: {token_resp.text}"}

    token = token_resp.json().get('access_token')
    hdrs_json = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    # ── 2. Resolver driveId + itemId ──────────────────────────────────────
    encoded = _b64.b64encode(SHAREPOINT_URL_WK.encode()).decode().rstrip('=')
    encoded = 'u!' + encoded.replace('/', '_').replace('+', '-')
    res = requests.get(
        f'https://graph.microsoft.com/v1.0/shares/{encoded}/driveItem',
        headers=hdrs_json, timeout=20,
    )
    if res.status_code != 200:
        return {"ok": False, "error": f"No se pudo resolver el archivo: {res.text}"}

    item     = res.json()
    drive_id = item.get('parentReference', {}).get('driveId')
    item_id  = item.get('id')
    if not drive_id or not item_id:
        return {"ok": False, "error": "No se pudo obtener driveId o itemId."}

    wb_url = f'https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/workbook'

    # ── 3. Abrir sesión de workbook (permite trabajar con archivo abierto) ─
    sess_resp = requests.post(
        f'{wb_url}/createSession',
        headers=hdrs_json,
        json={"persistChanges": True},
        timeout=30,
    )
    if sess_resp.status_code not in (200, 201):
        return {"ok": False, "error": f"Error abriendo sesión: {sess_resp.text}"}

    session_id = sess_resp.json().get('id')
    hdrs = {**hdrs_json, "workbook-session-id": session_id}

    try:
        # ── 4. Verificar que la hoja no exista ────────────────────────────
        sheets_resp = requests.get(f'{wb_url}/worksheets', headers=hdrs, timeout=20)
        if sheets_resp.status_code != 200:
            return {"ok": False, "error": f"Error listando hojas: {sheets_resp.text}"}
        nombres = [h['name'].strip() for h in sheets_resp.json().get('value', [])]
        if nombre_hoja.upper() in [n.upper() for n in nombres]:
            return {"ok": False, "error": f"La hoja '{nombre_hoja}' ya existe."}

        # ── 5. Crear la hoja nueva ────────────────────────────────────────
        add_resp = requests.post(
            f'{wb_url}/worksheets/add',
            headers=hdrs,
            json={"name": nombre_hoja},
            timeout=20,
        )
        if add_resp.status_code not in (200, 201):
            return {"ok": False, "error": f"Error creando hoja: {add_resp.text}"}

        # ── 6. Mover la hoja al inicio (posición 0) ───────────────────────
        ws_id = add_resp.json().get('id', nombre_hoja)
        # Graph API usa el nombre como id en la URL
        move_resp = requests.patch(
            f'{wb_url}/worksheets/{nombre_hoja}',
            headers=hdrs,
            json={"position": 0},
            timeout=20,
        )
        # No es fatal si falla el reordenamiento

        # ── 7. Escribir celdas en lotes (batchUpdate vía range) ───────────
        #   Graph API permite escribir un rango completo de una vez con:
        #   PATCH /workbook/worksheets/{id}/range(address='A1:Z200')
        #   Body: { "values": [[row0col0, row0col1, ...], [row1col0, ...]] }
        #
        #   Construimos una matriz 175 x 19 (filas 1-175, cols A-S)
        NROWS, NCOLS = 175, 19  # cols A(0)..S(18)
        col_idx = {c: i for i, c in enumerate("ABCDEFGHIJKLMNOPQRS")}
        matrix = [[""] * NCOLS for _ in range(NROWS)]

        for cell in _celdas_de_la_hoja(nombre_hoja):
            addr = cell["address"]          # ej "B3"
            val  = cell["value"]
            # Parsear dirección
            col_str = ''.join(ch for ch in addr if ch.isalpha())
            row_str = ''.join(ch for ch in addr if ch.isdigit())
            if col_str in col_idx and row_str:
                r = int(row_str) - 1
                col_c = col_idx[col_str]
                if 0 <= r < NROWS and 0 <= col_c < NCOLS:
                    matrix[r][col_c] = val if val is not None else ""

        range_addr = f"A1:S{NROWS}"
        patch_resp = requests.patch(
            f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{range_addr}\')',
            headers=hdrs,
            json={"values": matrix},
            timeout=60,
        )
        if patch_resp.status_code not in (200, 201):
            return {"ok": False, "error": f"Error escribiendo celdas: {patch_resp.text}"}

        # ── 8. Aplicar formatos via Graph API ─────────────────────────────
        def fmt(rng, body):
            requests.patch(
                f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{rng}\')/format',
                headers=hdrs, json=body, timeout=30,
            )

        def fill(rng, color):
            requests.patch(
                f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{rng}\')/format/fill',
                headers=hdrs, json={"color": f"#{color}"}, timeout=30,
            )

        def font(rng, bold=False, color=None, size=None):
            body = {"bold": bold, "name": "Arial"}
            if color: body["color"] = f"#{color}"
            if size:  body["size"]  = size
            requests.patch(
                f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{rng}\')/format/font',
                headers=hdrs, json=body, timeout=30,
            )

        def border(rng, left=None, right=None, top=None, bottom=None):
            style_map = {"thin": "Continuous", "medium": "Medium"}
            base = f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{rng}\')/format/borders'
            for side_name, style in [("EdgeLeft",left),("EdgeRight",right),("EdgeTop",top),("EdgeBottom",bottom)]:
                if style:
                    requests.patch(
                        f'{base}/{side_name}',
                        headers=hdrs,
                        json={"style": style_map.get(style, style), "color": "#000000"},
                        timeout=30,
                    )

        # ── Colores de fondo ──────────────────────────────────────────────
        fill("B2",        "DAE3F3")   # azul claro encabezado semana
        fill("B3",        "C5E0B4")   # verde lima código semana

        # Verde claro (CCFFCC) — columnas USD L:S
        for rng in [
            "L5:S9",
            "L10:S21", "L22:S23", "L24:S60",
            "L61:S62", "L63:S71", "L72:S74",
            "L94:S100", "L102:S102", "L104:S104", "L107:S107",
        ]:
            fill(rng, "CCFFCC")

        # Naranja (FFCC99) — subtotales
        fill("C22:J22",   "FFCC99")
        fill("L22:S22",   "FFCC99")
        fill("C61:J61",   "FFCC99")
        fill("L61:S61",   "FFCC99")
        fill("C72:J72",   "FFCC99")
        fill("L72:S72",   "FFCC99")
        fill("B101:J101", "FFCC99")
        fill("L101:S101", "FFCC99")

        # Amarillo claro (FFFFCC) — producción y $/Ha
        fill("L76:S92",   "FFFFCC")
        fill("L110:S121", "FFFFCC")
        # Amarillo vivo (FFFF00) — charolas/esquejes MXN
        fill("D89:J91",   "FFFF00")

        # Verde oscuro (008000) — headers KPI
        for rng in ["B125", "L125", "B143", "L143", "N143", "B165"]:
            fill(rng, "008000")

        # Blanco explícito — sección KPI proyectos / logística (filas 126-172)
        fill("B126:J172", "FFFFFF")
        fill("L126:S172", "FFFFFF")

        # ── Color de texto navy (#333399) en todo el cuerpo + tamaño 10 ──
        font("B1:J175",  bold=False, color="333399", size=10)
        font("L1:S175",  bold=False, color="333399", size=10)

        # ── Negritas ──────────────────────────────────────────────────────
        font("B1:B3",    bold=True,  color="333399", size=10)
        font("B9",       bold=True,  color="333399", size=10)
        font("B22",      bold=True,  color="333399", size=10)
        font("B61",      bold=True,  color="333399", size=10)
        font("B72",      bold=True,  color="333399", size=10)
        font("B74",      bold=True,  color="333399", size=10)
        font("B93:B97",  bold=True,  color="333399", size=10)
        font("B101",     bold=True,  color="333399", size=10)
        font("B103",     bold=True,  color="333399", size=10)
        font("B105",     bold=True,  color="333399", size=10)
        font("B106",     bold=True,  color="333399", size=10)
        font("B108",     bold=True,  color="333399", size=10)
        font("B110",     bold=True,  color="333399", size=10)
        font("B116",     bold=True,  color="333399", size=10)
        font("B118",     bold=True,  color="333399", size=10)
        font("B119",     bold=True,  color="333399", size=10)
        font("B121",     bold=True,  color="333399", size=10)
        font("B124",     bold=True,  color="333399", size=10)
        font("B140",     bold=True,  color="333399", size=10)
        # Columna C subtotales / columna L subtotales USD
        for rng in ["C22:J22", "C61:J61", "C72:J72", "C74:J74",
                    "L22:S22", "L61:S61", "L72:S72", "L74:S74",
                    "L76:L92", "L95:L121",
                    "L101:S101", "L103:S103", "L105:S106", "L108:S108",
                    "L111:N114", "L116:N119", "L121:S121"]:
            font(rng, bold=True, size=10)
        # Columna C negrita en todas las filas de datos
        for rng in ["C10:C21", "C24:C60", "C63:C70",
                    "C76:C92", "C95:C121"]:
            font(rng, bold=True, color="333399", size=10)
        # KPI headers — texto blanco negrita
        for rng in ["B125", "L125", "B143", "L143", "N143", "B165"]:
            font(rng, bold=True, color="FFFFFF", size=10)
        # Texto azul en valores KPI proyectos/logística
        for rng in ["C126:C172", "L126:L172"]:
            font(rng, bold=False, color="0000FF", size=10)

        # ── Bordes — estrategia simplificada (pocas llamadas) ─────────────
        # ESTRUCTURA PRINCIPAL: 3 columnas clave con rangos grandes
        # Left medio en B (toda el área de datos)
        border("B2:B175",  left="medium")
        # Right medio en J (toda el área de datos)
        border("J2:J175",  right="medium")
        # Right thin en C (separador columna TOTAL)
        border("C5:C175",  right="thin")
        # Left medio en L + right medio en S (todo el bloque USD)
        border("L5:L175",  left="medium")
        border("S5:S175",  right="medium")
        # Left medio en L para separar C de la zona MXN izquierda también
        border("C5:C9",    left="medium")
        border("C10:C21",  left="medium")
        border("C22:C74",  left="medium")
        border("C76:C121", left="medium")

        # FILAS ESPECIALES — separadores horizontales MXN
        border("B5:J5",    top="medium", bottom="thin")
        border("B9:J9",    bottom="thin")
        border("B22:J22",  top="thin",   bottom="thin")
        border("B61:J61",  top="thin",   bottom="thin")
        border("B72:J72",  top="thin",   bottom="thin")
        border("B74:J74",  top="thin",   bottom="medium")
        border("B76:J76",  top="medium")
        border("B92:J92",  bottom="medium")
        border("B94:J94",  top="medium")
        border("B108:J108",bottom="medium")
        border("B110:J110",top="medium")
        border("B121:J121",bottom="medium")

        # FILAS ESPECIALES — separadores horizontales USD
        border("L5:S5",    top="medium", bottom="thin")
        border("L9:S9",    bottom="thin")
        border("L10:S10",  top="thin")
        border("L22:S22",  top="thin",   bottom="thin")
        border("L61:S61",  top="thin",   bottom="thin")
        border("L72:S72",  top="thin",   bottom="thin")
        border("L74:S74",  top="thin",   bottom="medium")
        border("L76:S76",  top="medium")
        border("L92:S92",  bottom="medium")
        border("L94:S94",  top="medium")
        border("L97:S97",  bottom="thin")
        border("L98:S98",  top="thin")
        border("L100:S100",bottom="thin")
        border("L101:S101",bottom="thin")
        border("L103:S103",top="thin",   bottom="thin")
        border("L105:S105",top="thin")
        border("L106:S106",bottom="thin")
        border("L108:S108",top="thin",   bottom="medium")
        border("L110:S110",top="medium")
        border("L113:S113",bottom="thin")
        border("L114:S114",top="thin",   bottom="thin")
        border("L116:S116",top="thin",   bottom="thin")
        border("L118:S118",top="thin")
        border("L119:S119",bottom="thin")
        border("L121:S121",top="thin",   bottom="medium")

        # KPI HEADERS borders
        border("B125",  left="thin", right="thin", top="thin")
        border("L125",  left="thin", right="thin", top="thin", bottom="thin")
        border("B143",  left="thin", right="thin", top="thin")
        border("J143",  right="thin", top="thin")
        border("L143",  left="thin", right="thin", top="thin", bottom="thin")
        border("N143",  left="thin", right="thin", top="thin", bottom="thin")
        border("B165",  left="thin", right="thin", top="thin")

        # PROYECTOS — outline
        border("B126:J139", left="thin",   right="thin")
        border("B139:J139", bottom="thin")
        border("L126:S139", left="thin",   right="thin")
        border("L139:S139", bottom="thin")
        border("B140",  left="thin", right="thin", top="thin", bottom="thin")
        border("L140",  left="thin", right="thin", bottom="thin")

        # LOGÍSTICA — outline
        border("B144:B150", left="thin")
        border("J144:J150", right="thin")
        border("L144:L150", left="thin")
        border("N144:N150", right="thin")
        border("S144:S150", right="thin")

        # ── Alineación ────────────────────────────────────────────────────
        fmt("B2",    {"horizontalAlignment": "Center"})
        fmt("B3",    {"horizontalAlignment": "Center"})
        fmt("B4",    {"horizontalAlignment": "Center"})
        fmt("C5:J5", {"horizontalAlignment": "Center"})
        fmt("L5:S5", {"horizontalAlignment": "Center", "verticalAlignment": "Center"})
        fmt("B6",    {"horizontalAlignment": "Center", "verticalAlignment": "Top", "wrapText": True})
        fmt("B7",    {"horizontalAlignment": "Center", "verticalAlignment": "Top", "wrapText": True})
        fmt("C7:J7", {"horizontalAlignment": "Center", "verticalAlignment": "Top"})
        fmt("L7:S7", {"horizontalAlignment": "Center", "verticalAlignment": "Top"})
        fmt("C8",    {"horizontalAlignment": "Center"})
        fmt("L8",    {"horizontalAlignment": "Center"})
        fmt("B9",    {"horizontalAlignment": "Center"})
        fmt("L125",  {"horizontalAlignment": "Center"})
        fmt("L143",  {"horizontalAlignment": "Center"})
        fmt("N143",  {"horizontalAlignment": "Center"})

        # ── Anchos de columnas ────────────────────────────────────────────
        # Configurar anchos de columna para que coincidan con el formato esperado
        # Graph API requiere usar el endpoint de columnas específico
        column_widths = {
            "A": 3,
            "B": 69.4,
            "C": 14,
            "D": 11, "E": 11, "F": 11, "G": 11, "H": 11, "I": 11, "J": 11,
            "K": 3,
            "L": 11, "M": 11, "N": 11, "O": 11, "P": 11, "Q": 11, "R": 11, "S": 11,
        }
        for col_letter, width in column_widths.items():
            try:
                # Usamos el endpoint de formato de rango para establecer el ancho
                requests.patch(
                    f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{col_letter}:{col_letter}\')/format',
                    headers=hdrs,
                    json={"columnWidth": width * 7.5},
                    timeout=20,
                )
            except Exception as e:
                # No es crítico si falla el ajuste de ancho
                print(f"⚠️  Error configurando ancho columna {col_letter}: {e}")

        # ── Alto de filas ─────────────────────────────────────────────────
        # Configurar alto de filas específicas
        row_heights = {
            3: 15.0,
            4: 15.0,
            6: 26.4,
        }
        for row_num, height in row_heights.items():
            try:
                requests.patch(
                    f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{row_num}:{row_num}\')/format',
                    headers=hdrs,
                    json={"rowHeight": height},
                    timeout=20,
                )
            except Exception as e:
                # No es crítico si falla el ajuste de alto
                print(f"⚠️  Error configurando alto fila {row_num}: {e}")

        # ── Formato de número (#,##0) para celdas de valores ─────────────
        # Aplicar formato de número con separador de miles a las celdas con valores
        number_ranges = [
            # Subtotales MXN
            "C22:J22", "C61:J61", "C72:J72", "C74:J74",
            # Subtotales USD
            "L22:S22", "L61:S61", "L72:S72", "L74:S74",
            # Valores de datos MXN
            "C10:J21", "C24:J60", "C63:J70",
            # Valores de datos USD
            "L10:S21", "L24:S60", "L63:S70",
            # Sección de producción y costos
            "C76:J92", "L76:S92",
            "C95:J121", "L95:S121",
        ]
        for rng in number_ranges:
            try:
                requests.patch(
                    f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{rng}\')/format',
                    headers=hdrs,
                    json={"numberFormat": "#,##0"},
                    timeout=20,
                )
            except Exception as e:
                print(f"⚠️  Error configurando formato número en {rng}: {e}")

        # ── Merge de celdas ───────────────────────────────────────────────
        merges = [
            # Headers principales
            "C5:J5", "L5:R5",
            # Headers columnas K-L
            "K1:L1", "K2:L2", "K4:L4",
            # Separadores K-L
            "K75:L75", "K93:L93", "K109:L109",
            # KPI headers K-L
            "K122:L122", "K123:L123", "K124:L124",
            "K141:L141", "K142:L142",
            "K164:L164",
            "K174:L174", "K175:L175",
            # KPI sections A-B
            "A123:B123",
            "A141:B141", "A142:B142",
            "A164:B164",
            "A174:B174", "A175:B175",
            # Valores combinados verticalmente (logística)
            "C153:C154", "L153:L154",
            "C157:C158", "L157:L158",
            "C161:C162", "L161:L162",
            "C167:C168", "L167:L168",
            "C171:C172", "L171:L172",
        ]
        for m in merges:
            try:
                requests.post(
                    f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{m}\')/merge',
                    headers=hdrs, json={"across": False}, timeout=20,
                )
            except Exception as e:
                print(f"⚠️  Error merge {m}: {e}")

    finally:
        # ── 8. Cerrar la sesión siempre ───────────────────────────────────
        requests.post(
            f'{wb_url}/closeSession',
            headers=hdrs,
            timeout=20,
        )

    return {
        "ok": True,
        "mensaje": f"Hoja '{nombre_hoja}' creada exitosamente en SharePoint.",
    }


# ─── Descarga de una hoja WK#### como xlsx con formato completo ───────────────
def get_sheet_xlsx(week_code: str) -> bytes | None:
    """
    Descarga el Excel de SharePoint y extrae la hoja WK{week_code}
    como un archivo .xlsx independiente con formato completo.
    """
    archivo = _descargar_excel(SHAREPOINT_URL_WK, "Excel WK")
    if archivo is None:
        return None

    archivo_bytes = archivo.getvalue()
    sheet_name = f"WK{week_code}"

    try:
        wb = openpyxl.load_workbook(BytesIO(archivo_bytes))

        target = None
        for sname in wb.sheetnames:
            normalized = re.sub(r'\s+', '', sname.strip()).upper()
            if normalized == sheet_name.upper():
                target = sname
                break

        if target is None:
            return None

        src_ws = wb[target]
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = target

        for row in src_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font          = copy(cell.font)
                    new_cell.border        = copy(cell.border)
                    new_cell.fill          = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection    = copy(cell.protection)
                    new_cell.alignment     = copy(cell.alignment)

        for merge in src_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merge))

        for col_letter, col_dim in src_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width  = col_dim.width
            new_ws.column_dimensions[col_letter].hidden = col_dim.hidden
        for row_num, row_dim in src_ws.row_dimensions.items():
            new_ws.row_dimensions[row_num].height = row_dim.height
            new_ws.row_dimensions[row_num].hidden = row_dim.hidden

        buf = BytesIO()
        new_wb.save(buf)
        buf.seek(0)
        return buf.read()

    except Exception as e:
        print(f"⚠️  Error extrayendo hoja {sheet_name}: {e}")
        return None
